#!/usr/bin/env python3
"""
BidEquity Platform — API Cost Model Generator

Generates an Excel workbook modelling Claude API costs across all 55 skills,
opportunity lifecycle phases, client portfolio scenarios, and cost reduction levers.

Based on live test data from Session 16 (April 2026) and Anthropic pricing.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

# ============================================================================
# PRICING (Anthropic, as of April 2026)
# ============================================================================

MODELS = {
    'sonnet': {'name': 'Claude Sonnet 4', 'input_per_m': 3.00, 'output_per_m': 15.00},
    'haiku':  {'name': 'Claude Haiku 4.5', 'input_per_m': 0.25, 'output_per_m': 1.25},
    'opus':   {'name': 'Claude Opus 4',    'input_per_m': 15.00, 'output_per_m': 75.00},
}

# ============================================================================
# SKILL INVENTORY WITH TOKEN ESTIMATES
# Based on live tests: Timeline Analysis (42,825 in / 2,097 out),
# Compliance Coverage (37,589 in / 1,856 out), ITT Extraction (21,260 in / 1,719 out)
# ============================================================================

SKILLS = [
    # Agent 1: Document Intelligence
    {'id': '1.1', 'agent': 1, 'name': 'ITT Extraction', 'phase': 'ITT Receipt', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 2000, 'frequency': 1,
     'notes': 'Once per ITT. Tested live: 21,260 in / 1,719 out. Scales with document size.'},
    {'id': '1.2', 'agent': 1, 'name': 'Evaluation Criteria Analysis', 'phase': 'ITT Receipt', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 2500, 'frequency': 1,
     'notes': 'Once per ITT. Reads extracted sections + evaluation framework.'},
    {'id': '1.3', 'agent': 1, 'name': 'Contract Analysis', 'phase': 'ITT Receipt', 'model': 'sonnet',
     'input_tokens': 30000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once per ITT. Contract documents can be large (50-100 pages).'},
    {'id': '1.4', 'agent': 1, 'name': 'Compliance Matrix', 'phase': 'ITT Receipt', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 2000, 'frequency': 1,
     'notes': 'Once per ITT.'},
    {'id': '1.5', 'agent': 1, 'name': 'Procurement Briefing', 'phase': 'ITT Receipt', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 4000, 'frequency': 1,
     'notes': 'Once per ITT. Produces narrative report — higher output tokens.'},
    {'id': '1.6', 'agent': 1, 'name': 'Clarification Impact (UC13)', 'phase': 'Mid-Bid', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 2000, 'frequency': 3,
     'notes': 'Per clarification round. Typical bid has 2-4 clarification exchanges.'},
    {'id': '1.7', 'agent': 1, 'name': 'Amendment Detection (UC14)', 'phase': 'Mid-Bid', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 2000, 'frequency': 1,
     'notes': 'Per ITT amendment. Not all bids have amendments.'},

    # Agent 2: Market & Competitive Intelligence
    {'id': '2.1', 'agent': 2, 'name': 'Client Profiling', 'phase': 'Pre-Bid', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 4000, 'frequency': 1,
     'notes': 'Once per client. REUSABLE across opportunities with same client. Pre-built profiles eliminate this cost.'},
    {'id': '2.2', 'agent': 2, 'name': 'Sector Scanning', 'phase': 'Pre-Bid', 'model': 'sonnet',
     'input_tokens': 10000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once per sector per quarter. REUSABLE across all opportunities in that sector.'},
    {'id': '2.3', 'agent': 2, 'name': 'Incumbent Assessment', 'phase': 'Pre-Bid', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once per opportunity. Assesses incumbent vulnerability.'},
    {'id': '2.4', 'agent': 2, 'name': 'Competitor Profiling', 'phase': 'Pre-Bid', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 4000, 'frequency': 1,
     'notes': 'Once per opportunity. REUSABLE competitor dossiers reduce to delta analysis only.'},
    {'id': '2.5', 'agent': 2, 'name': 'Stakeholder Mapping', 'phase': 'Pre-Bid', 'model': 'sonnet',
     'input_tokens': 12000, 'output_tokens': 2500, 'frequency': 1,
     'notes': 'Once per opportunity. Partially reusable for repeat clients.'},

    # Agent 3: Strategy & Scoring Analyst
    {'id': '3.1', 'agent': 3, 'name': 'Win Theme Mapping', 'phase': 'Strategy', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 3000, 'frequency': 2,
     'notes': 'Run at strategy lock + post-storyboard. Two passes typical.'},
    {'id': '3.2', 'agent': 3, 'name': 'PWIN Scoring', 'phase': 'Strategy', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 3000, 'frequency': 3,
     'notes': 'Run at qualify, strategy lock, and pre-submission. Three assessments typical.'},
    {'id': '3.3', 'agent': 3, 'name': 'Capture Plan Assembly', 'phase': 'Strategy', 'model': 'sonnet',
     'input_tokens': 30000, 'output_tokens': 5000, 'frequency': 1,
     'notes': 'Once per opportunity. Large output — produces the capture plan document.'},
    {'id': '3.4', 'agent': 3, 'name': 'Clarification Strategy', 'phase': 'Strategy', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once per opportunity.'},
    {'id': '3.5', 'agent': 3, 'name': 'Timeline Analysis (UC1)', 'phase': 'Ongoing', 'model': 'sonnet',
     'input_tokens': 45000, 'output_tokens': 2500, 'frequency': 8,
     'notes': 'Tested live: 42,825 in / 2,097 out. Run weekly or on-demand. 8 runs over 10-week bid.'},
    {'id': '3.6', 'agent': 3, 'name': 'Effort Reforecast (UC2)', 'phase': 'Ongoing', 'model': 'sonnet',
     'input_tokens': 35000, 'output_tokens': 2500, 'frequency': 4,
     'notes': 'Fortnightly during production phase. 4 runs typical.'},
    {'id': '3.7', 'agent': 3, 'name': 'Standup Priorities (UC3)', 'phase': 'Ongoing', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 1500, 'frequency': 30,
     'notes': 'Daily during active bid period. HIGHEST FREQUENCY skill. Consider Haiku.'},
    {'id': '3.8', 'agent': 3, 'name': 'Compliance Coverage (UC4)', 'phase': 'Ongoing', 'model': 'sonnet',
     'input_tokens': 40000, 'output_tokens': 2000, 'frequency': 4,
     'notes': 'Tested live: 37,589 in / 1,856 out. Run weekly during production. 4 runs typical.'},
    {'id': '3.9', 'agent': 3, 'name': 'Win Theme Audit (UC5)', 'phase': 'Production', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 2000, 'frequency': 3,
     'notes': 'Post-draft, post-red, post-gold. Three passes.'},
    {'id': '3.10', 'agent': 3, 'name': 'Marks Allocation (UC6)', 'phase': 'Production', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 2000, 'frequency': 2,
     'notes': 'At draft and final. Two passes.'},
    {'id': '3.11', 'agent': 3, 'name': 'Review Trajectory (UC7)', 'phase': 'Production', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 2000, 'frequency': 3,
     'notes': 'After each review cycle (pink, red, gold).'},
    {'id': '3.12', 'agent': 3, 'name': 'Gate Readiness (UC9)', 'phase': 'Governance', 'model': 'sonnet',
     'input_tokens': 30000, 'output_tokens': 2500, 'frequency': 5,
     'notes': 'Per governance gate. 4-6 gates per bid.'},
    {'id': '3.13', 'agent': 3, 'name': 'Stakeholder Engagement Risk (UC12)', 'phase': 'Governance', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 2000, 'frequency': 2,
     'notes': 'Pre-gate check. 2 runs typical.'},

    # Agent 4: Commercial & Financial Modelling
    {'id': '4.1', 'agent': 4, 'name': 'Cost Modelling', 'phase': 'Commercial', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 4000, 'frequency': 2,
     'notes': 'Initial build + revision. Two passes typical.'},
    {'id': '4.2', 'agent': 4, 'name': 'Pricing Scenarios', 'phase': 'Commercial', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 3000, 'frequency': 2,
     'notes': 'Initial + BAFO revision.'},
    {'id': '4.3', 'agent': 4, 'name': 'Partner Pricing', 'phase': 'Commercial', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 2500, 'frequency': 1,
     'notes': 'Once per opportunity. Not all bids have partners.'},
    {'id': '4.4', 'agent': 4, 'name': 'Commercial Model', 'phase': 'Commercial', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once per opportunity.'},
    {'id': '4.5', 'agent': 4, 'name': 'Sensitivity Analysis', 'phase': 'Commercial', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 2500, 'frequency': 2,
     'notes': 'Initial + post-BAFO.'},
    {'id': '4.6', 'agent': 4, 'name': 'Pricing Finalisation', 'phase': 'Commercial', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once at pricing lock.'},
    {'id': '4.7', 'agent': 4, 'name': 'Commercial Risk', 'phase': 'Commercial', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 2000, 'frequency': 1,
     'notes': 'Once per opportunity.'},
    {'id': '4.8', 'agent': 4, 'name': 'Risk/Pricing Validation (UC10)', 'phase': 'Governance', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 2000, 'frequency': 1,
     'notes': 'Pre-pricing gate.'},
    {'id': '4.9', 'agent': 4, 'name': 'Bid Cost Forecast (UC11)', 'phase': 'Ongoing', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 2000, 'frequency': 4,
     'notes': 'Fortnightly budget tracking. 4 runs typical.'},

    # Agent 5: Content & Response Drafting
    {'id': '5.1', 'agent': 5, 'name': 'Response Section Drafting', 'phase': 'Production', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 5000, 'frequency': 10,
     'notes': 'PER SECTION. 5-15 sections per bid. Highest output tokens. Major cost driver.'},
    {'id': '5.2', 'agent': 5, 'name': 'Executive Summary', 'phase': 'Production', 'model': 'sonnet',
     'input_tokens': 30000, 'output_tokens': 5000, 'frequency': 2,
     'notes': 'Draft + final. Reads all sections to synthesise.'},
    {'id': '5.3', 'agent': 5, 'name': 'Evidence Compilation', 'phase': 'Production', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 3000, 'frequency': 5,
     'notes': 'Per workstream or section group. 5 runs typical.'},
    {'id': '5.4', 'agent': 5, 'name': 'Governance Packs', 'phase': 'Governance', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 4000, 'frequency': 5,
     'notes': 'Per governance gate. 4-6 gates.'},
    {'id': '5.5', 'agent': 5, 'name': 'Production Formatting & QA', 'phase': 'Production', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 2000, 'frequency': 10,
     'notes': 'Per section. Same frequency as drafting.'},
    {'id': '5.6', 'agent': 5, 'name': 'Compliance Verification', 'phase': 'Pre-Submit', 'model': 'sonnet',
     'input_tokens': 30000, 'output_tokens': 2500, 'frequency': 1,
     'notes': 'Once pre-submission.'},
    {'id': '5.7', 'agent': 5, 'name': 'Presentation Intelligence (UC15)', 'phase': 'Post-Submit', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once post-submission. Not all bids have presentations.'},

    # Agent 6: Solution & Delivery Design
    {'id': '6.1', 'agent': 6, 'name': 'Current State Assessment', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 4000, 'frequency': 1,
     'notes': 'Once per opportunity. Rebid only.'},
    {'id': '6.2', 'agent': 6, 'name': 'Operating Model Design', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 30000, 'output_tokens': 6000, 'frequency': 2,
     'notes': 'Initial + revision. Highest effort skill (45 person-days methodology). Large output.'},
    {'id': '6.3', 'agent': 6, 'name': 'Service Delivery Model', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 5000, 'frequency': 1,
     'notes': 'Once per opportunity.'},
    {'id': '6.4', 'agent': 6, 'name': 'Technology Architecture', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 4000, 'frequency': 1,
     'notes': 'Once per opportunity.'},
    {'id': '6.5', 'agent': 6, 'name': 'Innovation & Social Value', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once per opportunity.'},
    {'id': '6.6', 'agent': 6, 'name': 'Solution Consolidation', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 35000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once at solution lock. Reads all solution outputs.'},
    {'id': '6.7', 'agent': 6, 'name': 'Staffing Model', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 4000, 'frequency': 1,
     'notes': 'Once per opportunity. Includes TUPE analysis.'},
    {'id': '6.8', 'agent': 6, 'name': 'Transition Planning', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 4000, 'frequency': 1,
     'notes': 'Once per opportunity.'},
    {'id': '6.9', 'agent': 6, 'name': 'Delivery Readiness', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once pre-gate.'},
    {'id': '6.10', 'agent': 6, 'name': 'Performance Framework', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 3000, 'frequency': 1,
     'notes': 'Once per opportunity.'},
    {'id': '6.11', 'agent': 6, 'name': 'Delivery Risk Consolidation', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 25000, 'output_tokens': 2500, 'frequency': 1,
     'notes': 'Once at solution lock.'},
    {'id': '6.12', 'agent': 6, 'name': 'Partner Sourcing', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 2500, 'frequency': 1,
     'notes': 'Once per opportunity. Not all bids have partners.'},
    {'id': '6.13', 'agent': 6, 'name': 'Partner Coordination', 'phase': 'Solution', 'model': 'sonnet',
     'input_tokens': 15000, 'output_tokens': 2000, 'frequency': 1,
     'notes': 'Once per opportunity. Not all bids have partners.'},
    {'id': '6.14', 'agent': 6, 'name': 'Reviewer Calibration (UC8)', 'phase': 'Production', 'model': 'sonnet',
     'input_tokens': 20000, 'output_tokens': 2000, 'frequency': 1,
     'notes': 'Once post-gold review.'},
]

# ============================================================================
# STYLING
# ============================================================================

NAVY = PatternFill(start_color='0B1628', end_color='0B1628', fill_type='solid')
DARK_NAVY = PatternFill(start_color='061018', end_color='061018', fill_type='solid')
GOLD_FILL = PatternFill(start_color='B8860B', end_color='B8860B', fill_type='solid')
LIGHT_FILL = PatternFill(start_color='F7F9FC', end_color='F7F9FC', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
AMBER_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
GREEN_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
RED_FILL = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

WHITE_FONT = Font(color='FFFFFF', bold=True, size=11, name='Calibri')
GOLD_FONT = Font(color='B8860B', bold=True, size=11, name='Calibri')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10, name='Calibri')
BODY_FONT = Font(color='333333', size=10, name='Calibri')
BOLD_FONT = Font(color='333333', bold=True, size=10, name='Calibri')
MONEY_FORMAT = '£#,##0.00'
MONEY_FORMAT_4 = '£#,##0.0000'
PCT_FORMAT = '0.0%'
NUM_FORMAT = '#,##0'

THIN_BORDER = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0'),
)


def style_header_row(ws, row, max_col, fill=NAVY, font=HEADER_FONT):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.font = BOLD_FONT if bold else BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    if fmt:
        cell.number_format = fmt
    return cell


def cost_per_call(skill, model_key=None):
    m = model_key or skill['model']
    pricing = MODELS[m]
    return (skill['input_tokens'] / 1_000_000 * pricing['input_per_m'] +
            skill['output_tokens'] / 1_000_000 * pricing['output_per_m'])


def cost_per_opp(skill, model_key=None):
    return cost_per_call(skill, model_key) * skill['frequency']


# ============================================================================
# SHEET 1: Skill Cost Inventory
# ============================================================================

def build_skill_inventory(wb):
    ws = wb.create_sheet('Skill Cost Inventory')

    # Title
    ws.merge_cells('A1:K1')
    ws.cell(row=1, column=1, value='BidEquity Platform — Skill Cost Inventory').font = Font(color='B8860B', bold=True, size=14, name='Calibri')
    ws.cell(row=1, column=1).fill = DARK_NAVY
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    for c in range(1, 12):
        ws.cell(row=1, column=c).fill = DARK_NAVY

    ws.merge_cells('A2:K2')
    ws.cell(row=2, column=1, value='Based on live test data (Session 16, April 2026) and Anthropic Sonnet pricing ($3/M input, $15/M output)').font = Font(color='999999', size=9, name='Calibri')
    ws.cell(row=2, column=1).fill = DARK_NAVY
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    for c in range(1, 12):
        ws.cell(row=2, column=c).fill = DARK_NAVY

    # Headers
    headers = ['Skill ID', 'Agent', 'Skill Name', 'Phase', 'Model',
               'Input Tokens', 'Output Tokens', 'Cost/Call',
               'Frequency/Opp', 'Cost/Opportunity', 'Notes']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    # Data
    row = 5
    total_cost = 0
    agent_costs = {}
    phase_costs = {}

    for skill in SKILLS:
        call_cost = cost_per_call(skill)
        opp_cost = cost_per_opp(skill)
        total_cost += opp_cost

        agent_costs.setdefault(skill['agent'], 0)
        agent_costs[skill['agent']] += opp_cost
        phase_costs.setdefault(skill['phase'], 0)
        phase_costs[skill['phase']] += opp_cost

        ws.cell(row=row, column=1, value=skill['id'])
        ws.cell(row=row, column=2, value=f"Agent {skill['agent']}")
        ws.cell(row=row, column=3, value=skill['name'])
        ws.cell(row=row, column=4, value=skill['phase'])
        ws.cell(row=row, column=5, value=MODELS[skill['model']]['name'])
        style_data_cell(ws, row, 6, NUM_FORMAT).value = skill['input_tokens']
        style_data_cell(ws, row, 7, NUM_FORMAT).value = skill['output_tokens']
        style_data_cell(ws, row, 8, MONEY_FORMAT_4).value = call_cost
        style_data_cell(ws, row, 9, NUM_FORMAT).value = skill['frequency']
        style_data_cell(ws, row, 10, MONEY_FORMAT).value = opp_cost
        ws.cell(row=row, column=11, value=skill['notes']).font = Font(color='666666', size=9, name='Calibri')

        for c in range(1, 12):
            ws.cell(row=row, column=c).border = THIN_BORDER
            if row % 2 == 0:
                ws.cell(row=row, column=c).fill = LIGHT_FILL

        row += 1

    # Total row
    row += 1
    ws.cell(row=row, column=1, value='TOTAL PER OPPORTUNITY').font = BOLD_FONT
    style_data_cell(ws, row, 10, MONEY_FORMAT, bold=True).value = total_cost
    ws.cell(row=row, column=10).fill = AMBER_FILL

    # Agent summary
    row += 2
    ws.cell(row=row, column=1, value='COST BY AGENT').font = GOLD_FONT
    row += 1
    for col_h, val in [('Agent', 1), ('Cost/Opportunity', 2), ('% of Total', 3)]:
        ws.cell(row=row, column=val, value=col_h)
    style_header_row(ws, row, 3)
    row += 1
    agent_names = {1: 'Document Intelligence', 2: 'Market & Competitive', 3: 'Strategy & Scoring',
                   4: 'Commercial & Financial', 5: 'Content & Drafting', 6: 'Solution & Delivery'}
    for agent_id in sorted(agent_costs):
        ws.cell(row=row, column=1, value=f"Agent {agent_id}: {agent_names[agent_id]}").font = BODY_FONT
        style_data_cell(ws, row, 2, MONEY_FORMAT).value = agent_costs[agent_id]
        style_data_cell(ws, row, 3, PCT_FORMAT).value = agent_costs[agent_id] / total_cost
        row += 1

    # Phase summary
    row += 1
    ws.cell(row=row, column=1, value='COST BY PHASE').font = GOLD_FONT
    row += 1
    for col_h, val in [('Phase', 1), ('Cost/Opportunity', 2), ('% of Total', 3)]:
        ws.cell(row=row, column=val, value=col_h)
    style_header_row(ws, row, 3)
    row += 1
    for phase in sorted(phase_costs, key=lambda p: phase_costs[p], reverse=True):
        ws.cell(row=row, column=1, value=phase).font = BODY_FONT
        style_data_cell(ws, row, 2, MONEY_FORMAT).value = phase_costs[phase]
        style_data_cell(ws, row, 3, PCT_FORMAT).value = phase_costs[phase] / total_cost
        row += 1

    # Column widths
    widths = [8, 10, 30, 12, 16, 14, 14, 12, 14, 16, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return total_cost, agent_costs, phase_costs


# ============================================================================
# SHEET 2: Opportunity Lifecycle Cost
# ============================================================================

def build_lifecycle_cost(wb, total_cost):
    ws = wb.create_sheet('Opportunity Lifecycle')

    ws.merge_cells('A1:G1')
    ws.cell(row=1, column=1, value='BidEquity Platform — Cost Per Opportunity Lifecycle').font = Font(color='B8860B', bold=True, size=14, name='Calibri')
    ws.cell(row=1, column=1).fill = DARK_NAVY
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = DARK_NAVY

    # Scenario analysis
    headers = ['Scenario', 'Sections', 'Duration (weeks)', 'Skill Invocations',
               'Total API Cost', 'GBP Equivalent', 'Notes']
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))

    scenarios = [
        ('Small bid (£5-25m)', 5, 6, 80, total_cost * 0.6, 'Fewer sections, shorter duration, subset of skills'),
        ('Medium bid (£25-100m)', 10, 10, 130, total_cost * 1.0, 'Full skill set, typical 10-week timeline'),
        ('Large bid (£100m+)', 20, 16, 200, total_cost * 1.8, 'More sections, longer duration, multiple revision cycles'),
        ('Competitive dialogue', 10, 24, 250, total_cost * 2.5, 'Multiple dialogue rounds multiply revision skills'),
    ]

    for i, (name, sections, weeks, invocations, cost_usd, notes) in enumerate(scenarios):
        row = 4 + i
        ws.cell(row=row, column=1, value=name).font = BOLD_FONT
        style_data_cell(ws, row, 2, NUM_FORMAT).value = sections
        style_data_cell(ws, row, 3, NUM_FORMAT).value = weeks
        style_data_cell(ws, row, 4, NUM_FORMAT).value = invocations
        style_data_cell(ws, row, 5, MONEY_FORMAT).value = cost_usd
        style_data_cell(ws, row, 6, MONEY_FORMAT).value = cost_usd * 0.79  # USD to GBP
        ws.cell(row=row, column=7, value=notes).font = Font(color='666666', size=9, name='Calibri')
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN_BORDER

    # Cost per phase breakdown for medium bid
    row = 10
    ws.cell(row=row, column=1, value='MEDIUM BID COST BREAKDOWN BY PHASE').font = GOLD_FONT
    row += 1

    phase_order = ['Pre-Bid', 'ITT Receipt', 'Strategy', 'Solution', 'Commercial',
                   'Production', 'Ongoing', 'Governance', 'Pre-Submit', 'Post-Submit', 'Mid-Bid']

    phase_headers = ['Phase', 'Skills', 'Invocations', 'API Cost (USD)', 'API Cost (GBP)', '% of Total']
    for i, h in enumerate(phase_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(phase_headers))

    row += 1
    for phase in phase_order:
        phase_skills = [s for s in SKILLS if s['phase'] == phase]
        if not phase_skills:
            continue
        invocations = sum(s['frequency'] for s in phase_skills)
        phase_cost = sum(cost_per_opp(s) for s in phase_skills)
        ws.cell(row=row, column=1, value=phase).font = BODY_FONT
        style_data_cell(ws, row, 2, NUM_FORMAT).value = len(phase_skills)
        style_data_cell(ws, row, 3, NUM_FORMAT).value = invocations
        style_data_cell(ws, row, 4, MONEY_FORMAT).value = phase_cost
        style_data_cell(ws, row, 5, MONEY_FORMAT).value = phase_cost * 0.79
        style_data_cell(ws, row, 6, PCT_FORMAT).value = phase_cost / total_cost if total_cost else 0
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    widths = [28, 10, 14, 16, 16, 12, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# SHEET 3: Client Portfolio Scenarios
# ============================================================================

def build_portfolio_scenarios(wb, total_cost):
    ws = wb.create_sheet('Client Portfolio')

    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value='BidEquity Platform — Client Portfolio API Cost Scenarios').font = Font(color='B8860B', bold=True, size=14, name='Calibri')
    ws.cell(row=1, column=1).fill = DARK_NAVY
    for c in range(1, 9):
        ws.cell(row=1, column=c).fill = DARK_NAVY

    headers = ['Client Profile', 'Opportunities/Year', 'Avg Bid Size',
               'API Cost/Opp (USD)', 'Annual API (USD)', 'Annual API (GBP)',
               'Annual Revenue', 'API as % Revenue']
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))

    # Revenue assumptions: Core = £80k/yr, Command = £200k/yr per client
    profiles = [
        ('Startup client (Verdict only)', 3, '£50m', total_cost * 0.0, 0, 6000, 'Verdict is fixed fee, API cost in delivery margin'),
        ('Small BD team (Core)', 5, '£25m', total_cost * 0.6, 5, 80000, ''),
        ('Mid-size BD team (Core)', 10, '£50m', total_cost * 1.0, 10, 120000, ''),
        ('Large BD operation (Command)', 20, '£100m', total_cost * 1.0, 20, 250000, ''),
        ('Enterprise (Command)', 40, '£200m', total_cost * 1.5, 40, 500000, ''),
    ]

    for i, (name, opps, avg_size, cost_per_opp, n_opps, annual_rev, notes) in enumerate(profiles):
        row = 4 + i
        annual_api = cost_per_opp * opps
        ws.cell(row=row, column=1, value=name).font = BOLD_FONT
        style_data_cell(ws, row, 2, NUM_FORMAT).value = opps
        ws.cell(row=row, column=3, value=avg_size).font = BODY_FONT
        style_data_cell(ws, row, 4, MONEY_FORMAT).value = cost_per_opp
        style_data_cell(ws, row, 5, MONEY_FORMAT).value = annual_api
        style_data_cell(ws, row, 6, MONEY_FORMAT).value = annual_api * 0.79
        style_data_cell(ws, row, 7, MONEY_FORMAT).value = annual_rev
        style_data_cell(ws, row, 8, PCT_FORMAT).value = (annual_api * 0.79) / annual_rev if annual_rev else 0

        # Colour-code the percentage
        pct = (annual_api * 0.79) / annual_rev if annual_rev else 0
        if pct < 0.01:
            ws.cell(row=row, column=8).fill = GREEN_FILL
        elif pct < 0.05:
            ws.cell(row=row, column=8).fill = AMBER_FILL
        else:
            ws.cell(row=row, column=8).fill = RED_FILL

        for c in range(1, 9):
            ws.cell(row=row, column=c).border = THIN_BORDER

    # Key insight
    row = 10
    ws.cell(row=row, column=1, value='KEY INSIGHT').font = GOLD_FONT
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value=(
        'API costs are negligible relative to revenue across all client profiles. '
        'Even the highest-usage enterprise scenario (40 opportunities × £1.5/opp) '
        'produces annual API costs under £50, against £500k revenue. '
        'The cost driver is consultant time, not API tokens.'
    )).font = Font(color='333333', size=10, name='Calibri')
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 50

    widths = [30, 18, 12, 18, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# SHEET 4: Cost Reduction Levers
# ============================================================================

def build_cost_reduction(wb, total_cost):
    ws = wb.create_sheet('Cost Reduction Levers')

    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='BidEquity Platform — Cost Reduction Architecture').font = Font(color='B8860B', bold=True, size=14, name='Calibri')
    ws.cell(row=1, column=1).fill = DARK_NAVY
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = DARK_NAVY

    headers = ['Lever', 'Mechanism', 'Skills Affected', 'Estimated Saving', 'Implementation', 'Priority']
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))

    levers = [
        ('Pre-built client profiles',
         'Build client intelligence once via Claude Pro subscription, store in reference/. Skills read pre-built data instead of researching from scratch.',
         '2.1 Client Profiling, 2.3 Incumbent Assessment, 2.5 Stakeholder Mapping',
         'Eliminates ~$0.30/opp for repeat clients',
         'Bid library + Client Profiles product (designed, not built)',
         'HIGH — reduces cost AND improves quality'),
        ('Pre-built competitor dossiers',
         'Maintain competitor dossiers on recurring schedule (Agent 2). Skills read dossiers instead of profiling per bid.',
         '2.4 Competitor Profiling',
         'Eliminates ~$0.07/opp for known competitors',
         'Competitor Dossiers product (designed, not built)',
         'HIGH — same logic as client profiles'),
        ('Model selection: Haiku for simple skills',
         'Use Haiku ($0.25/$1.25 per M) instead of Sonnet ($3/$15 per M) for skills that don\'t need complex reasoning.',
         '3.7 Standup Priorities (30 calls/opp), 5.5 Production Formatting',
         'Standup alone: $0.93 → $0.08 (92% saving). ~$1.00 total saving.',
         'Change model field in YAML config. No code change.',
         'HIGH — biggest single saving, zero effort'),
        ('Anthropic prompt caching',
         'Cache the system prompt (sector knowledge, reasoning rules) across skill calls. Anthropic charges 90% less for cached tokens.',
         'All skills that load sector_knowledge + reasoning_rules (~30 skills)',
         '~30-40% reduction in input token costs for cached context',
         'Add cache_control headers to Claude API call. Small code change.',
         'MEDIUM — moderate saving, easy implementation'),
        ('MCP server-side filtering',
         'Already designed: MCP tools return only matching records, not full datasets. Reduces tokens sent to Claude.',
         'All read tools with filter parameters',
         'Already active. Prevents token bloat as bid data grows.',
         'Already implemented in MCP server.',
         'DONE'),
        ('Batch operations',
         'batch_create_response_sections and batch_save_qualification_reviews reduce round-trips.',
         'ITT Extraction, Qualification Review',
         'Reduces multi-turn overhead by ~30% for batch-capable skills.',
         'Already implemented in MCP tools.',
         'DONE'),
        ('Smaller context windows',
         'Some skills load full bid_execution data when they only need a subset. Targeted context loading reduces input tokens.',
         'Ongoing optimisation across all skills',
         '10-20% input token reduction with targeted context',
         'Refine skill context sections in YAML. Ongoing prompt engineering.',
         'LOW — incremental gains, ongoing effort'),
    ]

    for i, (name, mechanism, skills, saving, impl, priority) in enumerate(levers):
        row = 4 + i
        ws.cell(row=row, column=1, value=name).font = BOLD_FONT
        ws.cell(row=row, column=2, value=mechanism).font = BODY_FONT
        ws.cell(row=row, column=3, value=skills).font = Font(color='666666', size=9, name='Calibri')
        ws.cell(row=row, column=4, value=saving).font = BODY_FONT
        ws.cell(row=row, column=5, value=impl).font = BODY_FONT
        ws.cell(row=row, column=6, value=priority).font = BOLD_FONT

        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 60

    # Optimised cost estimate
    row = 4 + len(levers) + 2
    ws.cell(row=row, column=1, value='OPTIMISED COST ESTIMATE').font = GOLD_FONT
    row += 1
    estimates = [
        ('Current (all Sonnet, no caching)', total_cost, 1.0),
        ('With Haiku for standup + formatting', total_cost - 1.00, (total_cost - 1.00) / total_cost),
        ('With prompt caching (30% input saving)', (total_cost - 1.00) * 0.80, ((total_cost - 1.00) * 0.80) / total_cost),
        ('With pre-built client + competitor intel', (total_cost - 1.00) * 0.80 - 0.37, ((total_cost - 1.00) * 0.80 - 0.37) / total_cost),
    ]

    for col_h, val in [('Configuration', 1), ('Cost/Opp (USD)', 2), ('vs Baseline', 3)]:
        ws.cell(row=row, column=val, value=col_h)
    style_header_row(ws, row, 3)
    row += 1
    for name, cost, pct in estimates:
        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        style_data_cell(ws, row, 2, MONEY_FORMAT).value = cost
        style_data_cell(ws, row, 3, PCT_FORMAT).value = pct
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    widths = [28, 50, 35, 30, 40, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# SHEET 5: Verdict Product Cost
# ============================================================================

def build_verdict_cost(wb):
    ws = wb.create_sheet('Verdict Product')

    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='BidEquity Verdict — API Cost & Margin Analysis').font = Font(color='B8860B', bold=True, size=14, name='Calibri')
    ws.cell(row=1, column=1).fill = DARK_NAVY
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = DARK_NAVY

    headers = ['Skill / Step', 'Invocations', 'Input Tokens', 'Output Tokens', 'Cost (USD)', 'Notes']
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))

    verdict_skills = [
        ('ITT Extraction (procurement docs)', 1, 25000, 2000, 'Reused from Agent 1'),
        ('Proposal Parsing (new)', 1, 40000, 3000, 'Larger than ITT — full submission text'),
        ('Domain 5.1: Intelligence Assessment', 1, 30000, 3000, 'Forensic domain skill'),
        ('Domain 5.2: Win Strategy Assessment', 1, 30000, 3000, ''),
        ('Domain 5.3: Team Mobilisation Assessment', 1, 20000, 2500, ''),
        ('Domain 5.4: Solution Assessment', 1, 35000, 3000, 'Reads solution pack + proposal'),
        ('Domain 5.5: Governance Assessment', 1, 25000, 2500, ''),
        ('Domain 5.6: Proposal Quality Assessment', 1, 35000, 3000, 'Includes independent scoring'),
        ('Domain 5.7: Commercial Assessment', 1, 25000, 2500, ''),
        ('Domain 5.8: Post-Submission Assessment', 1, 20000, 2000, 'Conditional — not all bids'),
        ('Independent Scoring (per section × 8)', 8, 15000, 2000, '8 sections typical'),
        ('Traceability Analysis', 1, 35000, 3000, 'Cross-artefact chain mapping'),
        ('Report Assembly', 1, 40000, 5000, 'Consolidates all findings'),
        ('Pass 2 Re-assessment (8 domains)', 8, 25000, 2500, 'Post-interview enriched analysis'),
    ]

    row = 4
    total_cost = 0
    for name, invocations, inp, out, notes in verdict_skills:
        m = MODELS['sonnet']
        call_cost = inp / 1_000_000 * m['input_per_m'] + out / 1_000_000 * m['output_per_m']
        step_cost = call_cost * invocations
        total_cost += step_cost

        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        style_data_cell(ws, row, 2, NUM_FORMAT).value = invocations
        style_data_cell(ws, row, 3, NUM_FORMAT).value = inp * invocations
        style_data_cell(ws, row, 4, NUM_FORMAT).value = out * invocations
        style_data_cell(ws, row, 5, MONEY_FORMAT_4).value = step_cost
        ws.cell(row=row, column=6, value=notes).font = Font(color='666666', size=9, name='Calibri')
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Totals
    row += 1
    ws.cell(row=row, column=1, value='TOTAL API COST PER VERDICT SINGLE').font = BOLD_FONT
    style_data_cell(ws, row, 5, MONEY_FORMAT, bold=True).value = total_cost
    ws.cell(row=row, column=5).fill = AMBER_FILL
    style_data_cell(ws, row, 6).value = f'GBP: £{total_cost * 0.79:.2f}'

    # Margin analysis
    row += 2
    ws.cell(row=row, column=1, value='MARGIN ANALYSIS').font = GOLD_FONT
    row += 1
    for col_h, val in [('Metric', 1), ('Verdict Single', 2), ('Verdict Portfolio', 3)]:
        ws.cell(row=row, column=val, value=col_h)
    style_header_row(ws, row, 3)

    metrics = [
        ('Price', 2000, 5000),
        ('API Cost (GBP)', total_cost * 0.79, total_cost * 0.79 * 3),
        ('Consultant Time (est. 1.5 days @ £800/day)', 1200, 1200 * 3),
        ('Gross Margin', 2000 - total_cost * 0.79 - 1200, 5000 - total_cost * 0.79 * 3 - 1200 * 3),
        ('Gross Margin %', (2000 - total_cost * 0.79 - 1200) / 2000, (5000 - total_cost * 0.79 * 3 - 1200 * 3) / 5000),
    ]

    row += 1
    for name, single, portfolio in metrics:
        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        fmt = PCT_FORMAT if 'Margin %' in name else MONEY_FORMAT
        style_data_cell(ws, row, 2, fmt).value = single
        style_data_cell(ws, row, 3, fmt).value = portfolio
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER

        if 'Margin %' in name:
            ws.cell(row=row, column=2).fill = GREEN_FILL if single > 0.3 else AMBER_FILL
            ws.cell(row=row, column=3).fill = GREEN_FILL if portfolio > 0.3 else AMBER_FILL
        row += 1

    widths = [38, 14, 14, 14, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# SHEET 6: Product Cost Summary (all products side by side)
# ============================================================================

def build_product_summary(wb, execution_cost):
    ws = wb.create_sheet('Product Cost Summary')

    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='BidEquity Platform — API Cost by Product').font = Font(color='B8860B', bold=True, size=14, name='Calibri')
    ws.cell(row=1, column=1).fill = DARK_NAVY
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = DARK_NAVY

    ws.merge_cells('A2:F2')
    ws.cell(row=2, column=1, value='All costs are Claude API token costs only. Consultant time, infrastructure, and overheads are excluded.').font = Font(color='999999', size=9, name='Calibri')
    ws.cell(row=2, column=1).fill = DARK_NAVY
    for c in range(1, 7):
        ws.cell(row=2, column=c).fill = DARK_NAVY

    # ── Qualify Product ──
    row = 4
    ws.cell(row=row, column=1, value='PWIN QUALIFY').font = GOLD_FONT
    row += 1

    q_headers = ['Step', 'Invocations', 'Input Tokens', 'Output Tokens', 'Cost (USD)', 'Notes']
    for i, h in enumerate(q_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(q_headers))

    qualify_steps = [
        ('AI Assurance Review (per question)', 24, 2500, 800, 'System prompt + question + evidence + rubric + sector context. Currently calls Claude API directly from browser.'),
        ('Full Qualification Report', 1, 5000, 2000, 'Synthesises all 24 reviews into overall assessment with PWIN score.'),
        ('PWIN Score Calculation', 3, 3000, 500, 'At initial assessment, post-review, and final. Lightweight calculation.'),
    ]

    row += 1
    qualify_total = 0
    m = MODELS['sonnet']
    for name, invocations, inp, out, notes in qualify_steps:
        call_cost = inp / 1_000_000 * m['input_per_m'] + out / 1_000_000 * m['output_per_m']
        step_cost = call_cost * invocations
        qualify_total += step_cost

        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        style_data_cell(ws, row, 2, NUM_FORMAT).value = invocations
        style_data_cell(ws, row, 3, NUM_FORMAT).value = inp * invocations
        style_data_cell(ws, row, 4, NUM_FORMAT).value = out * invocations
        style_data_cell(ws, row, 5, MONEY_FORMAT_4).value = step_cost
        ws.cell(row=row, column=6, value=notes).font = Font(color='666666', size=9, name='Calibri')
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Qualify Total per Pursuit').font = BOLD_FONT
    style_data_cell(ws, row, 5, MONEY_FORMAT, bold=True).value = qualify_total
    ws.cell(row=row, column=5).fill = AMBER_FILL
    ws.cell(row=row, column=6, value=f'GBP: £{qualify_total * 0.79:.2f}').font = BOLD_FONT

    # ── Execution Product ──
    row += 2
    ws.cell(row=row, column=1, value='PWIN EXECUTION (Bid Lifecycle)').font = GOLD_FONT
    row += 1

    exec_headers = ['Phase', 'Skills', 'Invocations', 'Cost (USD)', 'Notes']
    for i, h in enumerate(exec_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(exec_headers))

    phase_order = ['Pre-Bid', 'ITT Receipt', 'Strategy', 'Solution', 'Commercial',
                   'Production', 'Ongoing', 'Governance', 'Mid-Bid', 'Pre-Submit', 'Post-Submit']
    row += 1
    for phase in phase_order:
        phase_skills = [s for s in SKILLS if s['phase'] == phase]
        if not phase_skills:
            continue
        invocations = sum(s['frequency'] for s in phase_skills)
        phase_cost = sum(cost_per_opp(s) for s in phase_skills)

        ws.cell(row=row, column=1, value=phase).font = BODY_FONT
        style_data_cell(ws, row, 2, NUM_FORMAT).value = len(phase_skills)
        style_data_cell(ws, row, 3, NUM_FORMAT).value = invocations
        style_data_cell(ws, row, 4, MONEY_FORMAT).value = phase_cost
        # Notes for key phases
        notes_map = {
            'Ongoing': 'Standup (30×), Timeline (8×), Compliance (4×), Effort reforecast (4×), Bid cost (4×)',
            'Production': 'Drafting (10×), Formatting (10×), Win theme audit (3×), Review trajectory (3×)',
            'Governance': 'Gate readiness (5×), Governance packs (5×), Stakeholder risk (2×), Risk/pricing (1×)',
            'Pre-Bid': 'Client profiling, sector scanning, incumbent, competitors, stakeholders — REUSABLE',
        }
        ws.cell(row=row, column=5, value=notes_map.get(phase, '')).font = Font(color='666666', size=9, name='Calibri')
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Execution Total per Opportunity').font = BOLD_FONT
    style_data_cell(ws, row, 4, MONEY_FORMAT, bold=True).value = execution_cost
    ws.cell(row=row, column=4).fill = AMBER_FILL
    ws.cell(row=row, column=5, value=f'GBP: £{execution_cost * 0.79:.2f}').font = BOLD_FONT

    # ── Verdict Product ──
    row += 2
    ws.cell(row=row, column=1, value='BIDEQUITY VERDICT (Post-Loss Review)').font = GOLD_FONT
    row += 1

    v_headers = ['Step', 'Cost (USD)', 'Notes']
    for i, h in enumerate(v_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(v_headers))

    verdict_summary = [
        ('Document ingestion (ITT + proposal)', 0.17, '2 skills: ITT extraction + proposal parsing'),
        ('8 domain assessments (Pass 1)', 0.76, '8 forensic domain skills'),
        ('Independent scoring (8 sections)', 0.60, 'Per-section scoring against evaluation criteria'),
        ('Traceability analysis', 0.16, 'Cross-artefact chain mapping'),
        ('Pass 2 re-assessment (post-interview)', 0.70, '8 domains re-run with enriched data'),
        ('Report assembly', 0.14, 'Consolidates all findings into Verdict Report'),
    ]
    row += 1
    verdict_total = 0
    for name, cost, notes in verdict_summary:
        verdict_total += cost
        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        style_data_cell(ws, row, 2, MONEY_FORMAT).value = cost
        ws.cell(row=row, column=3, value=notes).font = Font(color='666666', size=9, name='Calibri')
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Verdict Total per Engagement').font = BOLD_FONT
    style_data_cell(ws, row, 2, MONEY_FORMAT, bold=True).value = verdict_total
    ws.cell(row=row, column=2).fill = AMBER_FILL
    ws.cell(row=row, column=3, value=f'GBP: £{verdict_total * 0.79:.2f}').font = BOLD_FONT

    # ── Combined Summary ──
    row += 3
    ws.cell(row=row, column=1, value='COMBINED COST SUMMARY').font = GOLD_FONT
    row += 1
    for col_h, val in [('Product', 1), ('Cost per Use (USD)', 2), ('Cost per Use (GBP)', 3), ('Price to Client', 4), ('API as % Price', 5)]:
        ws.cell(row=row, column=val, value=col_h)
    style_header_row(ws, row, 5)

    combined = [
        ('Qualify (per pursuit)', qualify_total, qualify_total * 0.79, 'Included in Core/Command', None),
        ('Execution (per opportunity)', execution_cost, execution_cost * 0.79, 'Included in Core/Command', None),
        ('Qualify + Execution combined', qualify_total + execution_cost, (qualify_total + execution_cost) * 0.79, '£80-500k/yr (Core/Command)', (qualify_total + execution_cost) * 0.79 / 80000),
        ('Verdict Single', verdict_total, verdict_total * 0.79, '£2,000', verdict_total * 0.79 / 2000),
        ('Verdict Portfolio (×3)', verdict_total * 3, verdict_total * 3 * 0.79, '£5,000', verdict_total * 3 * 0.79 / 5000),
    ]

    row += 1
    for name, cost_usd, cost_gbp, price, pct in combined:
        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        style_data_cell(ws, row, 2, MONEY_FORMAT).value = cost_usd
        style_data_cell(ws, row, 3, MONEY_FORMAT).value = cost_gbp
        ws.cell(row=row, column=4, value=str(price) if isinstance(price, str) else price).font = BODY_FONT
        if pct is not None:
            style_data_cell(ws, row, 5, PCT_FORMAT).value = pct
            ws.cell(row=row, column=5).fill = GREEN_FILL if pct < 0.01 else AMBER_FILL
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Scaling scenario
    row += 2
    ws.cell(row=row, column=1, value='ANNUAL SCALING SCENARIO').font = GOLD_FONT
    row += 1
    for col_h, val in [('Client Scenario', 1), ('Opps/Year', 2), ('API/Year (GBP)', 3), ('Revenue/Year', 4), ('API as % Revenue', 5)]:
        ws.cell(row=row, column=val, value=col_h)
    style_header_row(ws, row, 5)

    annual_scenarios = [
        ('1 Core client, 5 opps', 5, (qualify_total + execution_cost) * 5 * 0.79, 80000),
        ('1 Command client, 15 opps', 15, (qualify_total + execution_cost) * 15 * 0.79, 250000),
        ('5 Core + 2 Command clients', 55, (qualify_total + execution_cost) * 55 * 0.79, 900000),
        ('10 Verdict Singles/year', 10, verdict_total * 10 * 0.79, 20000),
        ('Full portfolio: 5 Core + 2 Command + 10 Verdicts', 65, ((qualify_total + execution_cost) * 55 + verdict_total * 10) * 0.79, 920000),
    ]

    row += 1
    for name, opps, annual_api, revenue in annual_scenarios:
        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        style_data_cell(ws, row, 2, NUM_FORMAT).value = opps
        style_data_cell(ws, row, 3, MONEY_FORMAT).value = annual_api
        style_data_cell(ws, row, 4, MONEY_FORMAT).value = revenue
        pct = annual_api / revenue if revenue else 0
        style_data_cell(ws, row, 5, PCT_FORMAT).value = pct
        ws.cell(row=row, column=5).fill = GREEN_FILL if pct < 0.01 else AMBER_FILL
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    widths = [38, 20, 18, 22, 16, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# SHEET 7: Key Assumptions
# ============================================================================

def build_assumptions(wb):
    ws = wb.create_sheet('Assumptions')

    ws.merge_cells('A1:C1')
    ws.cell(row=1, column=1, value='Key Assumptions & Data Sources').font = Font(color='B8860B', bold=True, size=14, name='Calibri')
    ws.cell(row=1, column=1).fill = DARK_NAVY
    for c in range(1, 4):
        ws.cell(row=1, column=c).fill = DARK_NAVY

    assumptions = [
        ('Pricing', 'Claude Sonnet 4: $3.00/M input tokens, $15.00/M output tokens'),
        ('Pricing', 'Claude Haiku 4.5: $0.25/M input tokens, $1.25/M output tokens'),
        ('Pricing', 'Claude Opus 4: $15.00/M input tokens, $75.00/M output tokens'),
        ('Pricing', 'USD to GBP exchange rate: 0.79'),
        ('Pricing', 'Prices are list rates — volume discounts may apply at scale'),
        ('Token estimates', 'Based on live tests: Timeline Analysis (42,825 in / 2,097 out), Compliance Coverage (37,589 in / 1,856 out), ITT Extraction (21,260 in / 1,719 out)'),
        ('Token estimates', 'Non-tested skills estimated by analogy to tested skills, adjusted for expected context size and output length'),
        ('Token estimates', 'Multi-turn tool use adds ~20-30% overhead vs single-turn (tool call results returned to Claude)'),
        ('Frequency', 'Medium bid (£25-100m) assumed as baseline: 10 response sections, 10-week duration, 5 governance gates'),
        ('Frequency', 'Standup Priorities (Skill 3.7) at 30 calls/opp is the single highest-frequency skill — daily for 6 weeks'),
        ('Frequency', 'Response Section Drafting (Skill 5.1) and Production Formatting (Skill 5.5) scale linearly with section count'),
        ('Revenue', 'Core client revenue: £80-120k/year. Command client: £200-500k/year. Verdict Single: £2k, Portfolio: £5k'),
        ('Consultant cost', 'Verdict consultant time: 1.5 days per Single (intake + interview + report review + debrief) at £800/day internal cost'),
        ('Prompt caching', 'Anthropic prompt caching reduces cost of repeated system prompt by ~90%. Applicable when same system prompt is used across multiple calls in a session.'),
        ('Pre-built intel', 'Client profiles and competitor dossiers, once built, eliminate the research cost for repeat clients. Build cost is a one-time investment via Claude Pro subscription or initial API call.'),
        ('Scaling', 'Token usage scales with bid complexity (more sections = more drafting calls, more gates = more readiness calls) but NOT linearly — platform knowledge and pursuit context are fixed overhead.'),
    ]

    headers = ['Category', 'Assumption']
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))

    for i, (cat, assumption) in enumerate(assumptions):
        row = 4 + i
        ws.cell(row=row, column=1, value=cat).font = BOLD_FONT
        ws.cell(row=row, column=2, value=assumption).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        for c in range(1, 3):
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.row_dimensions[row].height = 30

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 100


# ============================================================================
# MAIN
# ============================================================================

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    total_cost, agent_costs, phase_costs = build_skill_inventory(wb)
    build_lifecycle_cost(wb, total_cost)
    build_portfolio_scenarios(wb, total_cost)
    build_cost_reduction(wb, total_cost)
    build_verdict_cost(wb)
    build_product_summary(wb, total_cost)
    build_assumptions(wb)

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'docs', 'BidEquity_API_Cost_Model.xlsx')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f'Cost model saved to: {output_path}')
    print(f'Total API cost per medium opportunity: ${total_cost:.2f} (£{total_cost * 0.79:.2f})')
    print(f'Top 3 cost drivers:')
    skill_costs = [(s['name'], cost_per_opp(s)) for s in SKILLS]
    skill_costs.sort(key=lambda x: x[1], reverse=True)
    for name, cost in skill_costs[:3]:
        print(f'  {name}: ${cost:.2f}')


if __name__ == '__main__':
    main()

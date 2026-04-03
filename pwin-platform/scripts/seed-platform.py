#!/usr/bin/env python3
"""
PWIN Platform — Seed Platform Knowledge Files

Extracts data from the two enrichment spreadsheets and writes JSON files
to ~/.pwin/platform/. Run once after cloning, or after spreadsheet updates.

Usage:
    python3 scripts/seed-platform.py

Requires: openpyxl (pip install openpyxl)

Source files:
    pwin-qualify/docs/PWIN_AI_Enrichment_Review.xlsx
    pwin-qualify/docs/BWIN Qualify_AI Design_Proforma_v2.xlsx

Output: 11 JSON files in ~/.pwin/platform/
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. Run: pip install openpyxl")
    sys.exit(1)

# Paths
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENRICHMENT = os.path.join(REPO_ROOT, "..", "pwin-qualify", "docs", "PWIN_AI_Enrichment_Review.xlsx")
PROFORMA = os.path.join(REPO_ROOT, "..", "pwin-qualify", "docs", "BWIN Qualify_AI Design_Proforma_v2.xlsx")
OUTPUT = os.path.join(os.path.expanduser("~"), ".pwin", "platform")


def clean(val):
    """Clean newlines from cell values."""
    if isinstance(val, str):
        return val.replace("\n", " ").strip()
    return val


def write_json(filename, data):
    path = os.path.join(OUTPUT, filename)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)
    count = len(data) if isinstance(data, (list, dict)) else "object"
    print(f"  {filename}: {count} entries")


def extract_enrichment():
    """Extract from PWIN_AI_Enrichment_Review.xlsx."""
    print(f"\nReading: {os.path.basename(ENRICHMENT)}")
    wb = openpyxl.load_workbook(ENRICHMENT, data_only=True)

    # Sheet 2: Sector Knowledge
    ws = wb["2. Sector Enrichment"]
    sectors = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1]:
            continue
        sectors[clean(row[1])] = {
            "keyTheme": clean(row[2] or ""),
            "promptBlock": clean(row[3] or ""),
        }
    write_json("sector_knowledge.json", sectors)

    # Sheet 3: Opportunity Types
    ws = wb["3. Opportunity Type"]
    opp_types = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1]:
            continue
        opp_types[clean(row[1])] = {
            "primaryEvaluationDriver": clean(row[2] or ""),
            "promptBlock": clean(row[3] or ""),
        }
    write_json("opportunity_types.json", opp_types)

    # Sheet 4: Sector x Opp Matrix
    ws = wb["4. Sector x Opp Matrix"]
    matrix = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1]:
            continue
        matrix.append({
            "sector": clean(row[1]),
            "opportunityType": clean(row[2] or ""),
            "intersectionIntelligence": clean(row[3] or ""),
        })
    write_json("sector_opp_matrix.json", matrix)

    # Sheet 5: Few-Shot Examples
    ws = wb["5. Few-Shot Examples"]
    examples = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1]:
            continue
        examples.append({
            "questionCategory": clean(row[1]),
            "verdict": clean(row[2] or ""),
            "evidenceSubmitted": clean(row[3] or ""),
            "idealAIResponse": clean(row[4] or ""),
        })
    write_json("few_shot_examples.json", examples)

    # Sheet 6: Output Schema
    ws = wb["6. Output Schema"]
    schema = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1]:
            continue
        schema.append({
            "field": clean(row[1]),
            "status": clean(row[2] or ""),
            "description": clean(row[3] or ""),
            "example": clean(row[4] or ""),
            "include": clean(row[6] or ""),
        })
    write_json("output_schema.json", schema)

    # Sheet 7: System Prompt Core
    ws = wb["7. System Prompt Core"]
    prompt = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1]:
            continue
        prompt.append({
            "section": clean(row[1]),
            "text": clean(row[2] or ""),
        })
    write_json("system_prompt.json", prompt)


def extract_proforma():
    """Extract from BWIN Qualify_AI Design_Proforma_v2.xlsx."""
    print(f"\nReading: {os.path.basename(PROFORMA)}")
    wb = openpyxl.load_workbook(PROFORMA, data_only=True)

    # Sheet 2: Data Points
    ws = wb["2. Data Points"]
    dp = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        dp.append({
            "id": clean(row[0] or ""),
            "section": clean(row[1] or ""),
            "label": clean(row[2] or ""),
            "definition": clean(row[3] or ""),
            "whyItMatters": clean(row[4] or ""),
            "primarySource": clean(row[5] or ""),
            "secondarySource": clean(row[6] or ""),
            "fallback": clean(row[7] or ""),
            "confidenceRatingRule": clean(row[10] or ""),
            "riskIfUnknown": clean(row[11] or ""),
        })
    write_json("data_points.json", dp)

    # Sheet 3: Reasoning Rules
    ws = wb["3. Reasoning Rules"]
    rules = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        rules.append({
            "id": clean(row[0] or ""),
            "category": clean(row[1] or ""),
            "ruleName": clean(row[2] or ""),
            "condition": clean(row[3] or ""),
            "then": clean(row[4] or ""),
            "riskLevel": clean(row[5] or ""),
            "overrideable": clean(row[6] or ""),
            "bipNotes": clean(row[7] or ""),
        })
    write_json("reasoning_rules.json", rules)

    # Sheet 5: Source Hierarchy
    ws = wb["5. Source Hierarchy"]
    sh = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        sh.append({
            "id": clean(row[0] or ""),
            "dataCategory": clean(row[1] or ""),
            "source1": clean(row[2] or ""),
            "source2": clean(row[3] or ""),
            "source3": clean(row[4] or ""),
            "bipAccess": clean(row[5] or ""),
            "notes": clean(row[6] or ""),
        })
    write_json("source_hierarchy.json", sh)

    # Sheet 6: Confidence Model
    ws = wb["6. Confidence Model"]
    ratings = []
    for row in ws.iter_rows(min_row=4, max_row=7, values_only=True):
        if not row[0]:
            continue
        ratings.append({
            "rating": clean(row[0] or ""),
            "label": clean(row[1] or ""),
            "definition": clean(row[2] or ""),
            "example": clean(row[3] or ""),
            "colour": clean(row[4] or ""),
            "userActionRequired": clean(row[5] or ""),
        })
    completeness = []
    for row in ws.iter_rows(min_row=11, max_row=19, values_only=True):
        if not row[0]:
            continue
        completeness.append({
            "section": clean(row[0] or ""),
            "fields": row[1],
            "weight": clean(row[2] or ""),
            "minimumToProceed": clean(row[3] or ""),
            "rationale": clean(row[4] or ""),
        })
    write_json("confidence_model.json", {"ratings": ratings, "completeness": completeness})

    # Sheet 7: Success Factors
    ws = wb["7. Success Factors"]
    sf = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        sf.append({
            "id": clean(row[0] or ""),
            "factor": clean(row[1] or ""),
            "description": clean(row[2] or ""),
            "priority": clean(row[3] or ""),
            "signOffOwner": clean(row[4] or ""),
        })
    write_json("success_factors.json", sf)


def main():
    print("PWIN Platform — Seeding Platform Knowledge")
    print(f"Output: {OUTPUT}")

    os.makedirs(OUTPUT, exist_ok=True)

    if not os.path.exists(ENRICHMENT):
        print(f"\nError: {ENRICHMENT} not found")
        sys.exit(1)
    if not os.path.exists(PROFORMA):
        print(f"\nError: {PROFORMA} not found")
        sys.exit(1)

    extract_enrichment()
    extract_proforma()

    files = [f for f in os.listdir(OUTPUT) if f.endswith(".json")]
    print(f"\nDone. {len(files)} files written to {OUTPUT}")


if __name__ == "__main__":
    main()

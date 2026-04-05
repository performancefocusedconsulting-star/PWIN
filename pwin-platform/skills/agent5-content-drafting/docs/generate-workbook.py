#!/usr/bin/env python3
"""Generate the Governance Pack Page Design Workbook as Excel."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# Colours matching BidEquity brand
NAVY = "021744"
TEAL = "5CA3B6"
PARCHMENT = "F7F4EE"
WHITE = "FFFFFF"
LIGHT_GREY = "E8E4DC"
MID_GREY = "D5D9E0"
GREEN = "2D8A5E"
AMBER = "C68A1D"
RED = "C15050"

# Fonts
hdr_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=NAVY)
col_hdr_font = Font(name="Calibri", size=10, bold=True, color=NAVY)
body_font = Font(name="Calibri", size=10, color="333333")
note_font = Font(name="Calibri", size=9, italic=True, color="666666")
question_font = Font(name="Calibri", size=10, color=NAVY)
ref_font = Font(name="Calibri", size=10, color="576482")

# Fills
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
parchment_fill = PatternFill(start_color=PARCHMENT, end_color=PARCHMENT, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

# Borders
thin_border = Border(
    left=Side(style="thin", color=MID_GREY),
    right=Side(style="thin", color=MID_GREY),
    top=Side(style="thin", color=MID_GREY),
    bottom=Side(style="thin", color=MID_GREY),
)

# Alignment
wrap = Alignment(wrap_text=True, vertical="top")
wrap_centre = Alignment(wrap_text=True, vertical="top", horizontal="center")

# Column definitions for page design sheets
COLUMNS = [
    ("Page #", 8),
    ("Page Title", 28),
    ("Page Type", 16),
    ("Layout", 22),
    ("Data Displayed", 45),
    ("Visualisation", 30),
    ("Data Source", 25),
    ("Notes", 35),
]

# Dropdown options
PAGE_TYPES = "Title,Dashboard,Narrative,Data Table,Tornado,RAG Matrix,Action Tracker,Mixed"
LAYOUTS = "Full-width,Two-column (50/50),Two-column (60/40),Two-column (40/60),Grid (2x2),Grid (3x1),Grid (4x1),Header + table,Header + cards"
VIS_OPTIONS = "Metric card,Table,Prose paragraph,Bullet list,RAG badge,RAG grid,Bar chart,Checklist,Status badges,Numbered list,Quote block"
DATA_SOURCES = "shared.json,qualify.json,bid_execution.json,win_strategy.json,gate_definitions,governance_history,workstream:commercial,workstream:legal,workstream:operations,workstream:technology,workstream:HR,workstream:MTT,workstream:risk,workstream:procurement"


def add_page_design_sheet(wb, title, short_name, num_rows, purpose, audience, design_questions, example_rows=None):
    """Create a page design sheet for one gate tier."""
    ws = wb.create_sheet(title=short_name)
    ws.sheet_properties.tabColor = TEAL

    # --- Title section ---
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = title
    c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    c.fill = navy_fill
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Purpose: {purpose}"
    c.font = note_font
    c.fill = parchment_fill
    c.alignment = wrap

    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value = f"Primary audience: {audience}"
    c.font = note_font
    c.fill = parchment_fill
    c.alignment = wrap
    ws.row_dimensions[3].height = 20

    # --- Column headers ---
    row = 5
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=row, column=col_idx, value=col_name)
        c.font = col_hdr_font
        c.fill = light_fill
        c.border = thin_border
        c.alignment = wrap_centre
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # --- Example rows (if provided) ---
    data_start = row + 1
    if example_rows:
        for r_idx, ex_row in enumerate(example_rows):
            r = data_start + r_idx
            ws.row_dimensions[r].height = 60
            for c_idx, val in enumerate(ex_row, 1):
                c = ws.cell(row=r, column=c_idx, value=val)
                c.font = body_font
                c.border = thin_border
                c.alignment = wrap
                if c_idx == 1:
                    c.alignment = wrap_centre
                c.fill = parchment_fill
        data_start += len(example_rows)

    # --- Empty rows for user input ---
    for r_idx in range(num_rows):
        r = data_start + r_idx
        ws.row_dimensions[r].height = 50
        for c_idx in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=r, column=c_idx)
            c.border = thin_border
            c.alignment = wrap
            c.font = body_font
            if c_idx == 1:
                c.alignment = wrap_centre
            # Alternate row fill
            if r_idx % 2 == 0:
                c.fill = white_fill
            else:
                c.fill = parchment_fill

    last_data_row = data_start + num_rows - 1

    # --- Data validations (dropdowns) ---
    dv_type = DataValidation(type="list", formula1=f'"{PAGE_TYPES}"', allow_blank=True)
    dv_type.error = "Pick from the list"
    dv_type.prompt = "Select page type"
    ws.add_data_validation(dv_type)
    dv_type.add(f"C{row+1}:C{last_data_row}")

    dv_layout = DataValidation(type="list", formula1=f'"{LAYOUTS}"', allow_blank=True)
    ws.add_data_validation(dv_layout)
    dv_layout.add(f"D{row+1}:D{last_data_row}")

    # --- Design questions section ---
    q_row = last_data_row + 3
    ws.merge_cells(f"A{q_row}:H{q_row}")
    c = ws.cell(row=q_row, column=1, value="Design Questions — Please Answer")
    c.font = section_font
    c.fill = light_fill

    for i, q in enumerate(design_questions, 1):
        r = q_row + i
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(row=r, column=1, value=f"Q{i}.")
        c.font = Font(name="Calibri", size=10, bold=True, color=TEAL)
        ws.merge_cells(f"C{r}:E{r}")
        c = ws.cell(row=r, column=3, value=q)
        c.font = question_font
        c.alignment = wrap
        ws.merge_cells(f"F{r}:H{r}")
        c = ws.cell(row=r, column=6, value="[Your answer here]")
        c.font = note_font
        c.alignment = wrap
        c.fill = parchment_fill
        c.border = thin_border
        ws.row_dimensions[r].height = 40

    return ws


def add_reference_sheet(wb):
    """Create reference sheet with available options."""
    ws = wb.create_sheet(title="Reference")
    ws.sheet_properties.tabColor = NAVY

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Reference — Available Page Types, Layouts & Visualisations"
    c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    c.fill = navy_fill
    ws.row_dimensions[1].height = 36

    # --- Page Types ---
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="Page Types")
    c.font = section_font
    c.fill = light_fill

    page_types = [
        ("Title", "Cover page, section dividers", "Pursuit name, gate, date, classification"),
        ("Dashboard", "Key metrics at a glance", "Grid of metric cards with big numbers"),
        ("Narrative", "Explanations, strategy, recommendations", "Prose with optional sidebar cards"),
        ("Data Table", "Financial data, structured comparisons", "P&L, PI assessment, sign-off sheet"),
        ("Tornado", "Risk sensitivity analysis", "Horizontal bars + scenario table"),
        ("RAG Matrix", "Status assessment grids", "Legal terms grid, compliance matrix"),
        ("Action Tracker", "Gate actions, outstanding items", "Table with status badges"),
        ("Mixed", "Combination of data + narrative", "Dashboard top + narrative bottom"),
    ]

    row += 1
    for h_idx, h in enumerate(["Type", "Best For", "Example"], 1):
        c = ws.cell(row=row, column=h_idx, value=h)
        c.font = col_hdr_font
        c.fill = light_fill
        c.border = thin_border
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50

    for pt in page_types:
        row += 1
        for c_idx, val in enumerate(pt, 1):
            c = ws.cell(row=row, column=c_idx, value=val)
            c.font = ref_font
            c.border = thin_border
            c.alignment = wrap

    # --- Layouts ---
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="Layouts")
    c.font = section_font
    c.fill = light_fill

    layouts = [
        ("Full-width", "Single column, edge to edge"),
        ("Two-column (50/50)", "Equal split"),
        ("Two-column (60/40)", "Main content left, supporting right"),
        ("Two-column (40/60)", "Supporting left, main content right"),
        ("Grid (2x2)", "Four equal quadrants"),
        ("Grid (3x1)", "Three equal columns"),
        ("Grid (4x1)", "Four equal columns (dashboard metrics)"),
        ("Header + table", "Title/summary top, full-width table below"),
        ("Header + cards", "Title/summary top, metric cards below"),
    ]

    row += 1
    for h_idx, h in enumerate(["Layout", "Description"], 1):
        c = ws.cell(row=row, column=h_idx, value=h)
        c.font = col_hdr_font
        c.fill = light_fill
        c.border = thin_border

    for lt in layouts:
        row += 1
        for c_idx, val in enumerate(lt, 1):
            c = ws.cell(row=row, column=c_idx, value=val)
            c.font = ref_font
            c.border = thin_border
            c.alignment = wrap

    # --- Visualisations ---
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="Visualisations")
    c.font = section_font
    c.fill = light_fill

    vis = [
        ("Metric card", "Single KPI", 'Big "£307m" with "Total Contract Value" label'),
        ("Table", "Structured multi-row data", "P&L, risk register, PI assessment"),
        ("Prose paragraph", "Context, narrative, strategy", "Executive summary, opportunity description"),
        ("Bullet list", "Key points, decisions required", "Numbered decisions for the board"),
        ("RAG badge", "Single status indicator", "Green dot next to a line item"),
        ("RAG grid", "Multi-item status matrix", "Legal terms with coloured cells"),
        ("Bar chart", "Comparative values", "Tornado risk bars"),
        ("Checklist", "Completion tracking", "Mandatory prerequisites"),
        ("Status badges", "Action/item status", '"Closed" / "Open" / "Overdue" pills'),
        ("Numbered list", "Ordered items, priorities", "Approval items sought"),
        ("Quote block", "Caveats, warnings, flags", '"[ASSERTION — no supporting evidence]"'),
    ]

    row += 1
    for h_idx, h in enumerate(["Visualisation", "Best For", "Example"], 1):
        c = ws.cell(row=row, column=h_idx, value=h)
        c.font = col_hdr_font
        c.fill = light_fill
        c.border = thin_border

    for v in vis:
        row += 1
        for c_idx, val in enumerate(v, 1):
            c = ws.cell(row=row, column=c_idx, value=val)
            c.font = ref_font
            c.border = thin_border
            c.alignment = wrap

    # --- Data Sources ---
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="Data Sources")
    c.font = section_font
    c.fill = light_fill

    sources = [
        ("shared.json", "Pursuit metadata", "Client, TCV, ACV, term, sector, deal type, dates"),
        ("qualify.json", "PWIN Qualify product data", "PWIN score, 6 category scores, AI assurance findings, recommendation"),
        ("bid_execution.json", "Bid Execution product data", "Activity status, risk register, resource allocation, bid cost, schedule"),
        ("win_strategy.json", "Win Strategy product data", "Win themes, competitive positioning, stakeholder map, buyer values"),
        ("gate_definitions", "Platform governance knowledge", "Gate prerequisites, decision criteria, IC trigger thresholds"),
        ("governance_history", "Previous gate records", "Past gate decisions, actions, conditions, dates"),
        ("workstream:commercial", "Commercial workstream docs", "Financial model, pricing strategy, payment mechanism, indexation"),
        ("workstream:legal", "Legal workstream docs", "Contract terms, mark-up strategy, liability, T&Cs grid"),
        ("workstream:operations", "Operations workstream docs", "Delivery solution, staffing model, PI achievability, TOM"),
        ("workstream:technology", "Technology workstream docs", "Systems architecture, integration, GDPR, cyber security"),
        ("workstream:HR", "HR workstream docs", "TUPE, pensions, redundancy, key personnel, union engagement"),
        ("workstream:MTT", "MTT workstream docs", "Transition plan, milestones, critical path, MTT team, LDs"),
        ("workstream:risk", "Risk workstream docs", "Risk register, tornado analysis, mitigation status"),
        ("workstream:procurement", "Procurement workstream docs", "Subcontractor status, supply chain, teaming agreements"),
    ]

    row += 1
    for h_idx, h in enumerate(["Source", "Description", "Contains"], 1):
        c = ws.cell(row=row, column=h_idx, value=h)
        c.font = col_hdr_font
        c.fill = light_fill
        c.border = thin_border

    for s in sources:
        row += 1
        for c_idx, val in enumerate(s, 1):
            c = ws.cell(row=row, column=c_idx, value=val)
            c.font = ref_font
            c.border = thin_border
            c.alignment = wrap


# ============================================================
# BUILD WORKBOOK
# ============================================================

# Remove default sheet
del wb["Sheet"]

# --- Qualification Tier ---
qual_example = [
    (1, "[Gate Name] — [Client]", "Title", "Full-width",
     "Pursuit name, client, gate, date, TCV headline, classification",
     "Centred display text", "shared.json", "Draft watermark if not final"),
    (2, "Deal Dashboard", "Dashboard", "Grid (4x1) + Grid (4x1)",
     "Row 1: Client, TCV, ACV, Term\nRow 2: Sector, Deal Type, Procurement Route, Deadline",
     "Metric cards", "shared.json", "IC triggers flagged in red if met"),
    (3, "Opportunity Summary", "Narrative", "Two-column (60/40)",
     "Left: what client is buying, why now, services in scope\nRight: key dates, procurement stage",
     "Prose + bullet list", "shared.json, qualify.json", "One page max"),
    (4, "Qualification Assessment", "Mixed", "Two-column (50/50)",
     "Left: PWIN score (large), 6 category scores as bars\nRight: AI assurance findings",
     "Metric card + bar chart + bullet list", "qualify.json", "Flag evidence gaps with quote blocks"),
    (5, "Competitive Landscape", "Narrative", "Full-width",
     "Known competitors, positioning, differentiators",
     "Prose + table", "win_strategy.json", "Conditional — only if intelligence exists"),
    (6, "Bid Cost & Resource", "Data Table", "Header + table",
     "Bid cost estimate, cost to Gate 4, key roles, timeline",
     "Table + metric cards", "bid_execution.json", ""),
    (7, "Recommendation", "Narrative", "Two-column (60/40)",
     "Left: recommendation, rationale, conditions\nRight: decisions required, approval sought",
     "Prose + numbered list", "qualify.json", "Always final page before any sign-off"),
]

add_page_design_sheet(
    wb, "Tier 1: Qualification Pack (Gates 0–2)", "1 Qualification",
    num_rows=10, example_rows=qual_example,
    purpose="Evidence-based pursue/no-pursue decision",
    audience="Divisional leadership / BLRT",
    design_questions=[
        "Should the PWIN Qualify assessment be one page or two? (scores summary + AI assurance findings separately?)",
        "Is there a 'decisions required' summary page at the end, or does the recommendation page cover this?",
        "Do you want an IC trigger assessment as its own page, or embedded in the dashboard?",
        "Any pages that should only appear conditionally? (e.g. competitive landscape only if intelligence exists)",
    ],
)

# --- Solution Tier ---
add_page_design_sheet(
    wb, "Tier 2: Solution Pack (Gate 3)", "2 Solution",
    num_rows=20,
    purpose="Confirm proposition post-ITT. Win strategy locked. Solution credible.",
    audience="BLRT + solution reviewers",
    design_questions=[
        "Should 'actions from previous gate' come before or after the deal dashboard?",
        "How detailed should the preliminary financial summary be? One page with key metrics, or a full P&L table?",
        "Win strategy — one page with themes + positioning, or separate pages?",
        "Is the evaluation criteria / score-to-win a table, a visual matrix, or narrative?",
        "Top 10 risks at Gate 3 — table format or individual risk cards?",
        "How many pages total do you expect for this tier?",
    ],
)

# --- Submission Tier ---
add_page_design_sheet(
    wb, "Tier 3: Submission Pack (Gate 4)", "3 Submission",
    num_rows=40,
    purpose="Full approval to submit binding tender. Complete commercial, legal, operational picture.",
    audience="Investment Committee / Group CRC",
    design_questions=[
        "Lead with purpose/agenda then dashboard (like Capita/Serco), or executive summary first?",
        "Financial section — how many pages? (Capita had 5: P&L, Balance Sheet, Cashflow, NPV, Sensitivity. Serco had 3.)",
        "Risk tornado — summary page + one per category (like Capita), or condensed to summary + top risks?",
        "Delivery solution — one page per sub-section (partners, people, tech, assets, MTT, social value), or combine some?",
        "Legal terms — single dense grid (like Serco IC), or one page per topic area?",
        "Performance indicators — one table for all PIs, or split operational vs system PIs?",
        "Price-to-win / competitive — one page or two?",
        "Where does the sign-off control sheet sit — last page, or after recommendation?",
        "Should there be an appendix section for full tornado breakdowns, full PI tables?",
        "How many pages total do you expect? (Capita CRC was ~40, Serco IC was ~40)",
    ],
)

# --- Contract Tier ---
add_page_design_sheet(
    wb, "Tier 4: Contract Pack (Gate 5)", "4 Contract",
    num_rows=15,
    purpose="Authority to sign contract. Delta from Gate 4.",
    audience="Same as Gate 4 plus contract signatories",
    design_questions=[
        "Should every page show Gate 4 position vs final position (delta view), or just present the final state?",
        "Transition readiness — one checklist page, or detailed with Gantt/timeline reference?",
        "Full updated sign-off sheet, or just the areas that re-signed?",
    ],
)

# --- Reference Sheet ---
add_reference_sheet(wb)

# Save
output = "/workspaces/PWIN/pwin-platform/skills/agent5-content-drafting/docs/Governance-Pack-Page-Design-Workbook.xlsx"
wb.save(output)
print(f"Saved: {output}")

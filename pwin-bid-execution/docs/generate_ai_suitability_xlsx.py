#!/usr/bin/env python3
"""Generate AI Suitability Assessment Excel from methodology gold standard and assessment data.

Reads methodology_gold_standard.md and ai_suitability_assessment.md, produces
ai_suitability_assessment.xlsx with 5 sheets.
"""

import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOCS = Path(__file__).resolve().parent
GOLD_STANDARD = DOCS / "methodology_gold_standard.md"
AI_ASSESSMENT = DOCS / "ai_suitability_assessment.md"
OUTPUT = DOCS / "ai_suitability_assessment.xlsx"

# ── Styling ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=9)
BOLD_FONT = Font(name="Calibri", size=9, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

RATING_FILL = {"High": HIGH_FILL, "Medium": MEDIUM_FILL, "Low": LOW_FILL}


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def style_body(ws, nrows, ncols, rating_col=None):
    for row in range(2, nrows + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
        if rating_col:
            rating = ws.cell(row=row, column=rating_col).value
            if rating in RATING_FILL:
                ws.cell(row=row, column=rating_col).fill = RATING_FILL[rating]


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Extract tasks from gold standard ────────────────────────────────────────

def extract_tasks():
    content = GOLD_STANDARD.read_text(encoding="utf-8")
    tasks = []
    lines = content.split("\n")
    for i, line in enumerate(lines):
        m = re.search(r"id: '([A-Z]+-\d+\.\d+\.\d+)'", line)
        if m:
            task_id = m.group(1)
            for j in range(i, min(i + 3, len(lines))):
                nm = re.search(r"name: '(.+)'", lines[j])
                if nm:
                    name = nm.group(1).replace("\\'", "'")
                    # Clean up names that have RACI data appended
                    raci_idx = name.find("', raci:")
                    if raci_idx > 0:
                        name = name[:raci_idx]
                    tasks.append((task_id, name))
                    break
    return tasks


# ── Extract detailed AI assessments from markdown ───────────────────────────

def extract_detailed_assessments():
    content = AI_ASSESSMENT.read_text(encoding="utf-8")
    data = {}
    pattern = (
        r'\|\s*((?:SAL|SOL|COM|LEG|DEL|SUP|BM|PRD|GOV|POST)-\d+\.\d+\.\d+)'
        r'\s+(.+?)\s*\|\s*\*\*(High|Medium|Low)\*\*\s*\|'
        r'\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(\d+)%\s*\|'
    )
    for m in re.finditer(pattern, content):
        task_id = m.group(1)
        data[task_id] = {
            "rating": m.group(3),
            "what_ai": m.group(4).strip().replace("**", ""),
            "inputs": m.group(5).strip(),
            "value": m.group(6).strip().replace("**", ""),
            "role": m.group(7).strip(),
            "time_reduction": int(m.group(8)),
        }
    return data


# ── Summary-level ratings per activity (from the summary tables) ────────────

# Format: activity_code -> (high_count, med_count, low_count, key_ai_value, biggest_blocker)

ACTIVITY_SUMMARIES = {
    # SOL-03 through SOL-12
    "SOL-03": (0, 5, 4, "AI drafts design options against requirements. Human designs.", "Design is creative human work"),
    "SOL-04": (0, 4, 2, "AI maps processes to requirements. Human designs processes.", "Process design is human expertise"),
    "SOL-05": (1, 4, 1, "AI assesses technology options, maps security requirements", "Build/buy decisions need organisational context"),
    "SOL-06": (1, 3, 1, "AI models workforce costs, maps TUPE terms", "TUPE data often incomplete"),
    "SOL-07": (0, 4, 2, "AI drafts phase plans from SOL outputs. Human validates", "Transition planning requires delivery experience"),
    "SOL-08": (1, 5, 2, "AI scans technology landscape, models productivity curve", "Innovation strategy is human creative work"),
    "SOL-09": (1, 3, 1, "AI analyses local area data, maps commitments to SV model", "Local knowledge and commitment design is human"),
    "SOL-10": (3, 2, 1, "AI searches evidence library, identifies gaps, matches evidence to sections", "Evidence library must be structured and accessible"),
    "SOL-11": (1, 2, 1, "AI runs coherence checks across solution components", "Design validation requires human judgement"),
    "SOL-12": (2, 2, 1, "AI consolidates risks from all activities automatically", "Risk assessment requires domain expertise"),
    # COM
    "COM-01": (2, 4, 1, "AI builds cost model from structured staffing/technology data", "Assumptions require commercial judgement"),
    "COM-02": (2, 3, 1, "AI analyses Contracts Finder data, benchmarks, models price envelope", "Competitor price intelligence is limited"),
    "COM-03": (1, 3, 2, "AI validates partner submissions against our model", "Negotiation is human"),
    "COM-04": (0, 4, 2, "AI maps ITT payment provisions. Human designs commercial approach", "Commercial design is strategic judgement"),
    "COM-05": (3, 1, 1, "AI runs all sensitivity analysis and stress tests automatically", "Assumptions and risk appetite are human"),
    "COM-06": (1, 2, 1, "AI reconciles cost model to pricing schedules, checks arithmetic", "Pricing decisions are human/governance"),
    "COM-07": (2, 2, 1, "AI consolidates risks, cross-references sensitivities", "Risk acceptance is human governance"),
    # LEG
    "LEG-01": (2, 1, 1, "AI analyses contract clauses, identifies non-standard terms", "Legal judgement on acceptability is human"),
    "LEG-02": (1, 2, 0, "AI maps risk allocation across contract categories", "Risk appetite decisions are human"),
    "LEG-03": (1, 1, 1, "AI extracts insurance requirements, compares to current coverage", "Procurement of new coverage is operational"),
    "LEG-04": (0, 2, 1, "AI analyses TUPE regulations against staffing plan", "Legal judgement on compliance is human"),
    "LEG-05": (1, 1, 1, "AI drafts DPIA from technology design", "DPIA validation requires legal sign-off"),
    "LEG-06": (1, 1, 1, "AI compares partner terms to prime contract for flow-down gaps", "Negotiation strategy is human"),
    # DEL
    "DEL-01": (1, 2, 1, "AI consolidates risks from all sources", "Risk judgement is human"),
    "DEL-02": (0, 3, 1, "AI drafts programme from SOL-07 transition plan", "Programme design requires delivery experience"),
    "DEL-03": (1, 1, 1, "AI maps ITT KPIs to delivery components", "Confidence assessment is human"),
    "DEL-04": (0, 2, 1, "AI maps resource needs from staffing model to timeline", "Resource judgement is human"),
    "DEL-05": (1, 1, 0, "AI drafts BC/DR and exit plan from solution design", "Planning is straightforward once design exists"),
    "DEL-06": (1, 1, 1, "AI consolidates all risk registers for governance", "Risk acceptance is human governance"),
    # SUP
    "SUP-01": (1, 1, 1, "AI searches market databases for matching partners", "Due diligence requires human assessment"),
    "SUP-02": (0, 2, 1, "AI reviews partner submissions for completeness", "Integration judgement is human"),
    "SUP-03": (0, 1, 2, "AI reviews standard terms. Negotiation is human", "Negotiation is fundamentally human"),
    "SUP-04": (0, 1, 1, "AI validates partner pricing against our model", "Facilitation/escalation is human"),
    "SUP-05": (1, 1, 0, "AI reviews partner evidence for quality/relevance", "Collection depends on partner responsiveness"),
    "SUP-06": (0, 1, 2, "AI compares flow-down alignment. Negotiation is human", "Negotiation is human"),
    # BM
    "BM-01": (0, 1, 1, "AI drafts programme from template. Human confirms", "Kickoff is a human leadership event"),
    "BM-02": (1, 0, 0, "AI drafts BMP from template + capture plan", "Standard document assembly"),
    "BM-03": (1, 0, 0, "AI configures repository from template. Human confirms", "Straightforward automation"),
    "BM-04": (0, 1, 1, "AI tracks from team roster data. Human manages changes", "Change management is human"),
    "BM-05": (1, 0, 0, "AI automates budget vs actual tracking", "Standard financial tracking"),
    "BM-06": (1, 1, 0, "AI generates progress reports from activity status data", "Report generation, not insight"),
    "BM-07": (1, 1, 0, "AI generates page budgets from marks analysis", "Quality standards need human calibration"),
    "BM-08": (0, 1, 0, "AI drafts comms plan. Human manages relationships", "Stakeholder management is human"),
    "BM-09": (0, 1, 1, "AI prepares briefing packs. Sessions are human", "Dialogue is fundamentally human interaction"),
    "BM-10": (1, 1, 1, "AI drafts storyboard structure from win themes + scoring", "Storyboard design requires creative judgement"),
    "BM-11": (0, 0, 1, "AI structures the debrief template", "Debrief is a human conversation"),
    "BM-12": (1, 1, 1, "AI analyses review scores, identifies patterns", "Pattern analysis is AI-suitable"),
    "BM-13": (1, 1, 0, "AI consolidates all workstream risks automatically", "Risk judgement is human"),
    "BM-14": (1, 1, 0, "AI tracks submissions, logs responses, distributes", "Operational tracking automation"),
    "BM-15": (1, 1, 0, "AI triages responses against requirements baseline", "Impact identification is systematic"),
    "BM-16": (0, 1, 1, "AI compares current intelligence to original strategy", "Strategic reassessment is human"),
    # PRD
    "PRD-01": (2, 0, 0, "AI maps requirements to responses automatically", "Requires structured requirements data"),
    "PRD-02": (1, 2, 0, "AI produces first drafts from storyboard + solution", "Quality of first draft depends on solution detail"),
    "PRD-03": (1, 1, 0, "AI matches evidence to sections from requirements matrix", "Evidence library must be structured"),
    "PRD-04": (1, 1, 0, "AI populates pricing schedules from COM-06 locked model", "Standard data mapping"),
    "PRD-05": (0, 1, 1, "AI pre-assesses storyboard against evaluation criteria", "Review judgement is human"),
    "PRD-06": (0, 1, 1, "AI simulates evaluator scoring against marking scheme", "Evaluator simulation requires calibration"),
    "PRD-07": (0, 0, 1, "Executive review is fundamentally human", "Governance is human"),
    "PRD-08": (1, 1, 0, "AI checks cross-references, word counts, formatting", "Standard QA automation"),
    "PRD-09": (0, 1, 1, "AI runs final compliance check against matrix", "Portal upload may require human"),
    # GOV
    "GOV-01": (1, 0, 1, "AI assembles review pack from upstream outputs", "Decision is human governance"),
    "GOV-02": (1, 0, 1, "AI assembles review pack", "Decision is human governance"),
    "GOV-03": (1, 0, 1, "AI assembles review pack", "Decision is human governance"),
    "GOV-04": (1, 0, 1, "AI assembles executive summary", "Decision is human governance"),
    "GOV-05": (0, 0, 1, "Formal sign-off is entirely human", "Governance is human"),
    "GOV-06": (1, 0, 1, "AI assembles legal review pack", "Decision is human governance"),
    # POST
    "POST-01": (0, 1, 1, "AI drafts presentation from submitted proposal", "Presentation design is creative"),
    "POST-02": (0, 0, 1, "AI simulates Q&A questions. Rehearsal is human", "Human practice"),
    "POST-03": (0, 0, 1, "Cannot be AI'd", "Client-facing presentation"),
    "POST-04": (1, 1, 0, "AI drafts clarification responses consistent with proposal", "Consistency checking is systematic"),
    "POST-05": (0, 2, 0, "AI models price adjustments using sensitivity data", "BAFO strategy is human judgement"),
    "POST-06": (0, 0, 1, "Governance is human", "Human governance"),
    "POST-07": (0, 0, 1, "Cannot be AI'd", "Negotiation is human"),
    "POST-08": (0, 1, 1, "AI assembles handover pack from all bid outputs", "Knowledge transfer requires human"),
}

# ── Activity effort data (from methodology_gold_standard.md) ──────────────

ACTIVITY_EFFORT = {
    # SAL
    "SAL-01": {"effortDays": 10, "teamSize": 1, "name": "Customer engagement & intelligence gathering"},
    "SAL-02": {"effortDays": 5, "teamSize": 1, "name": "Incumbent performance review"},
    "SAL-03": {"effortDays": 5, "teamSize": 1, "name": "Competitor analysis & positioning"},
    "SAL-04": {"effortDays": 5, "teamSize": 1, "name": "Win theme development & refinement"},
    "SAL-05": {"effortDays": 3, "teamSize": 1, "name": "Evaluation criteria mapping & scoring strategy"},
    "SAL-06": {"effortDays": 3, "teamSize": 1, "name": "Capture plan finalisation & win strategy lock"},
    "SAL-07": {"effortDays": 3, "teamSize": 2, "name": "Pre-submission clarification strategy & drafting"},
    "SAL-10": {"effortDays": 5, "teamSize": 1, "name": "Stakeholder relationship mapping & engagement"},
    # SOL
    "SOL-01": {"effortDays": 5, "teamSize": 2, "name": "Requirements analysis & interpretation"},
    "SOL-02": {"effortDays": 5, "teamSize": 1, "name": "Current operating model assessment"},
    "SOL-03": {"effortDays": 45, "teamSize": 3, "name": "Target operating model design"},
    "SOL-04": {"effortDays": 30, "teamSize": 3, "name": "Service delivery model design"},
    "SOL-05": {"effortDays": 20, "teamSize": 2, "name": "Technology & digital approach"},
    "SOL-06": {"effortDays": 8, "teamSize": 1, "name": "Staffing model & TUPE analysis"},
    "SOL-07": {"effortDays": 16, "teamSize": 2, "name": "Transition & mobilisation approach"},
    "SOL-08": {"effortDays": 5, "teamSize": 1, "name": "Innovation & continuous improvement"},
    "SOL-09": {"effortDays": 5, "teamSize": 1, "name": "Social value proposition"},
    "SOL-10": {"effortDays": 3, "teamSize": 1, "name": "Evidence strategy & case study identification"},
    "SOL-11": {"effortDays": 3, "teamSize": 1, "name": "Solution design lock"},
    "SOL-12": {"effortDays": 3, "teamSize": 1, "name": "Solution risk identification & analysis"},
    # COM
    "COM-01": {"effortDays": 8, "teamSize": 1, "name": "Should-cost model development"},
    "COM-02": {"effortDays": 3, "teamSize": 1, "name": "Price-to-win analysis"},
    "COM-03": {"effortDays": 5, "teamSize": 1, "name": "Subcontractor & supply chain pricing"},
    "COM-04": {"effortDays": 5, "teamSize": 1, "name": "Commercial model & payment mechanisms"},
    "COM-05": {"effortDays": 3, "teamSize": 1, "name": "Margin structure & sensitivity analysis"},
    "COM-06": {"effortDays": 5, "teamSize": 1, "name": "Pricing model finalisation"},
    "COM-07": {"effortDays": 3, "teamSize": 1, "name": "Commercial risk identification & analysis"},
    # LEG
    "LEG-01": {"effortDays": 8, "teamSize": 1, "name": "Contract review & markup"},
    "LEG-02": {"effortDays": 3, "teamSize": 1, "name": "Risk allocation analysis"},
    "LEG-03": {"effortDays": 3, "teamSize": 1, "name": "Insurance & liability review"},
    "LEG-04": {"effortDays": 5, "teamSize": 1, "name": "TUPE obligations assessment"},
    "LEG-05": {"effortDays": 5, "teamSize": 1, "name": "Data protection & security review"},
    "LEG-06": {"effortDays": 5, "teamSize": 1, "name": "Subcontractor terms review"},
    # DEL
    "DEL-01": {"effortDays": 8, "teamSize": 1, "name": "Delivery risk & assumptions register"},
    "DEL-02": {"effortDays": 5, "teamSize": 1, "name": "Mobilisation & implementation planning"},
    "DEL-03": {"effortDays": 5, "teamSize": 1, "name": "Performance framework & KPIs/SLAs"},
    "DEL-04": {"effortDays": 3, "teamSize": 1, "name": "Resource & capacity plan"},
    "DEL-05": {"effortDays": 3, "teamSize": 1, "name": "Business continuity & exit planning"},
    "DEL-06": {"effortDays": 5, "teamSize": 1, "name": "Risk mitigation & residual acceptance"},
    # SUP
    "SUP-01": {"effortDays": 5, "teamSize": 1, "name": "Partner identification & selection"},
    "SUP-02": {"effortDays": 10, "teamSize": 2, "name": "Partner solution inputs & design contribution"},
    "SUP-03": {"effortDays": 15, "teamSize": 1, "name": "Teaming agreement negotiation"},
    "SUP-04": {"effortDays": 8, "teamSize": 1, "name": "Partner pricing"},
    "SUP-05": {"effortDays": 5, "teamSize": 2, "name": "Partner credentials & references"},
    "SUP-06": {"effortDays": 5, "teamSize": 1, "name": "Back-to-back commercial terms"},
    # BM
    "BM-01": {"effortDays": 3, "teamSize": 1, "name": "Kickoff planning & execution"},
    "BM-02": {"effortDays": 2, "teamSize": 1, "name": "Bid management plan production"},
    "BM-03": {"effortDays": 1, "teamSize": 1, "name": "Document management & version control setup"},
    "BM-04": {"effortDays": 0, "teamSize": 1, "name": "Resource management & tracking"},
    "BM-05": {"effortDays": 0, "teamSize": 1, "name": "Bid cost management"},
    "BM-06": {"effortDays": 0, "teamSize": 1, "name": "Progress reporting"},
    "BM-07": {"effortDays": 3, "teamSize": 1, "name": "Quality management approach"},
    "BM-08": {"effortDays": 2, "teamSize": 1, "name": "Stakeholder & communications management"},
    "BM-09": {"effortDays": 0, "teamSize": 2, "name": "Competitive dialogue management"},
    "BM-10": {"effortDays": 5, "teamSize": 2, "name": "Storyboard development & sign-off"},
    "BM-11": {"effortDays": 1, "teamSize": 1, "name": "Hot debrief (post-submission)"},
    "BM-12": {"effortDays": 3, "teamSize": 1, "name": "Lessons learned & knowledge capture"},
    "BM-13": {"effortDays": 0, "teamSize": 1, "name": "Bid risk & assumptions register"},
    "BM-14": {"effortDays": 5, "teamSize": 1, "name": "Clarification submission & response management"},
    "BM-15": {"effortDays": 3, "teamSize": 2, "name": "Clarification impact analysis & workstream updates"},
    "BM-16": {"effortDays": 2, "teamSize": 1, "name": "Win strategy refresh (post-ITT)"},
    # PRD
    "PRD-01": {"effortDays": 3, "teamSize": 1, "name": "Compliance matrix & requirements mapping"},
    "PRD-02": {"effortDays": 60, "teamSize": 4, "name": "Section drafting & content assembly"},
    "PRD-03": {"effortDays": 16, "teamSize": 2, "name": "Evidence, case studies & CV assembly"},
    "PRD-04": {"effortDays": 5, "teamSize": 1, "name": "Pricing schedules & commercial response"},
    "PRD-05": {"effortDays": 3, "teamSize": 1, "name": "Pink review (storyboard/outline)"},
    "PRD-06": {"effortDays": 5, "teamSize": 1, "name": "Red review (full draft)"},
    "PRD-07": {"effortDays": 3, "teamSize": 1, "name": "Gold review (final/executive)"},
    "PRD-08": {"effortDays": 3, "teamSize": 1, "name": "Final QA & formatting"},
    "PRD-09": {"effortDays": 2, "teamSize": 1, "name": "Submission packaging & upload"},
    # GOV
    "GOV-01": {"effortDays": 2, "teamSize": 1, "name": "Pursuit approval"},
    "GOV-02": {"effortDays": 2, "teamSize": 1, "name": "Solution & strategy review"},
    "GOV-03": {"effortDays": 2, "teamSize": 1, "name": "Pricing & risk review"},
    "GOV-04": {"effortDays": 2, "teamSize": 1, "name": "Executive approval"},
    "GOV-05": {"effortDays": 1, "teamSize": 1, "name": "Final submission authority"},
    "GOV-06": {"effortDays": 2, "teamSize": 1, "name": "Legal & contractual review"},
    # POST
    "POST-01": {"effortDays": 10, "teamSize": 2, "name": "Presentation design & development"},
    "POST-02": {"effortDays": 3, "teamSize": 1, "name": "Presentation rehearsals & coaching"},
    "POST-03": {"effortDays": 1, "teamSize": 1, "name": "Presentation delivery"},
    "POST-04": {"effortDays": 5, "teamSize": 2, "name": "Post-submission clarification management"},
    "POST-05": {"effortDays": 10, "teamSize": 2, "name": "BAFO preparation & revised pricing"},
    "POST-06": {"effortDays": 2, "teamSize": 1, "name": "BAFO governance approval"},
    "POST-07": {"effortDays": 5, "teamSize": 1, "name": "Contract negotiation support"},
    "POST-08": {"effortDays": 3, "teamSize": 1, "name": "Award processing & mobilisation handover"},
}


# ── Role mappings by workstream/activity ────────────────────────────────────

ROLE_MAP = {
    "SAL": "Capture Lead",
    "SOL": "Solution Architect",
    "COM": "Commercial Lead",
    "LEG": "Legal Lead",
    "DEL": "Delivery Director",
    "SUP": "Supply Chain Lead",
    "BM": "Bid Manager",
    "PRD": "Bid Manager",
    "GOV": "Bid Manager",
    "POST": "Bid Manager",
}

ACTIVITY_ROLE_OVERRIDES = {
    "SAL-05": "Bid Manager",
    "SAL-06": "Bid Director",
    "SOL-05": "Technical Lead",
    "SOL-06": "HR Lead",
    "SOL-07": "Delivery Director",
    "BM-16": "Bid Director",
    "PRD-02": "Writers",
    "PRD-03": "Bid Coordinator",
    "PRD-04": "Commercial Lead",
    "PRD-05": "Reviewers",
    "PRD-06": "Reviewers",
    "PRD-07": "Bid Director",
    "PRD-08": "Bid Coordinator",
    "GOV-01": "Bid Director",
    "GOV-02": "Bid Director",
    "GOV-03": "Bid Director",
    "GOV-04": "Bid Director",
    "GOV-05": "Bid Director",
    "GOV-06": "Bid Director",
    "POST-01": "Bid Director",
    "POST-02": "Bid Director",
    "POST-03": "Bid Director",
    "POST-07": "Bid Director",
}

# ── Phase assignments ───────────────────────────────────────────────────────

PHASE_1_TASKS = {
    "SOL-01.1.1", "SOL-01.1.2", "SOL-01.1.3", "SOL-01.2.1", "SOL-01.2.2", "SOL-01.2.3",
    "SAL-05.1.1", "SAL-05.1.2", "SAL-05.1.3", "SAL-05.2.1", "SAL-05.2.2", "SAL-05.2.3",
    "SAL-05.2.4", "SAL-05.2.5",
    "SAL-06.1.2",
    "LEG-01.1.1", "LEG-01.1.2", "LEG-01.2.1", "LEG-01.2.2",
    "PRD-01.1.1", "PRD-01.1.2",
    "SAL-07.1.1", "SAL-07.2.1",
}

PHASE_2_TASKS = {
    "PRD-02.1.1", "PRD-02.1.2", "PRD-02.1.3",
    "BM-10.1.1", "BM-10.1.2", "BM-10.2.1",
    "GOV-01.1.1", "GOV-02.1.1", "GOV-03.1.1", "GOV-04.1.1", "GOV-06.1.1",
    "BM-06.1.1", "BM-06.1.2",
    "PRD-08.1.1", "PRD-08.1.2",
    "SAL-04.2.2",
    "SAL-06.2.4", "SAL-06.4.1",
}

PHASE_3_TASKS = {
    "SAL-01.1.1", "SAL-01.1.2", "SAL-01.1.3", "SAL-01.1.4",
    "SAL-01.2.1", "SAL-01.2.2", "SAL-01.2.3",
    "SAL-02.1.1", "SAL-02.1.2", "SAL-02.1.3", "SAL-02.2.1", "SAL-02.2.2",
    "SAL-02.3.1", "SAL-02.3.2", "SAL-02.3.3", "SAL-02.3.4", "SAL-02.3.5",
    "SAL-03.1.1", "SAL-03.1.2", "SAL-03.1.3", "SAL-03.2.1", "SAL-03.2.2",
    "SAL-10.1.1", "SAL-10.1.2", "SAL-10.1.3", "SAL-10.2.1", "SAL-10.2.2", "SAL-10.2.3",
    "SOL-02.1.1", "SOL-02.1.2", "SOL-02.1.3", "SOL-02.2.1", "SOL-02.2.2", "SOL-02.2.3",
    "SOL-10.1.1", "SOL-10.1.2", "SOL-10.1.3", "SOL-10.2.1", "SOL-10.2.2", "SOL-10.2.3",
}

PHASE_4_TASKS = {
    "COM-01.1.1", "COM-01.1.2", "COM-01.1.3", "COM-01.1.4",
    "COM-01.2.1", "COM-01.2.2", "COM-01.2.3",
    "COM-02.1.1", "COM-02.1.2", "COM-02.1.3",
    "COM-02.2.1", "COM-02.2.2", "COM-02.2.3",
    "COM-05.1.1", "COM-05.1.2", "COM-05.2.1", "COM-05.2.2", "COM-05.2.3",
    "SAL-02.2.1", "SAL-02.3.4",
}

PHASE_5_TASKS = {
    "BM-13.1.1", "BM-13.1.2",
    "SOL-12.1.1", "SOL-12.1.2", "SOL-12.2.1", "SOL-12.2.2", "SOL-12.2.3",
    "COM-07.1.1", "COM-07.1.2", "COM-07.2.1", "COM-07.2.2", "COM-07.2.3",
    "DEL-06.1.1", "DEL-06.1.2", "DEL-06.1.3",
    "BM-15.1.1", "BM-15.1.2",
}


def get_phase(task_id):
    if task_id in PHASE_1_TASKS:
        return "Phase 1"
    if task_id in PHASE_2_TASKS:
        return "Phase 2"
    if task_id in PHASE_3_TASKS:
        return "Phase 3"
    if task_id in PHASE_4_TASKS:
        return "Phase 4"
    if task_id in PHASE_5_TASKS:
        return "Phase 5"
    return "N/A"


# ── Compound Acceleration values ──────────────────────────────────────────
# Tasks that benefit from upstream AI outputs already being complete.
# First-in-chain tasks get 0%. Downstream tasks get higher values.

def get_compound_acceleration(task_id, activity_code, task_name):
    """Assign compound acceleration % based on upstream AI dependency."""
    name_lower = task_name.lower()
    ws = activity_code.split("-")[0]

    # First-in-chain tasks — no upstream acceleration
    if activity_code in ("SAL-01", "SOL-01", "SOL-02"):
        return 0

    # HIGH compound (25-30%) — downstream of document intelligence, massive prep elimination
    # Solution design tasks — benefit from SOL-01 requirements + SAL intelligence already done
    if activity_code in ("SOL-03", "SOL-04", "SOL-05"):
        return 25
    # Storyboard — benefits from win themes + eval criteria + solution all AI-prepared
    if activity_code == "BM-10":
        return 30
    # Response drafting — benefits from storyboard + solution + win themes + evidence all ready
    if activity_code == "PRD-02":
        return 30
    # Price-to-win — benefits from SOL cost data + competitive analysis already done
    if activity_code == "COM-02":
        return 25
    # Capture plan / win strategy lock — all upstream SAL intelligence available
    if activity_code == "SAL-06":
        return 25
    # Evidence assembly — benefits from evidence strategy + gap analysis already done
    if activity_code == "PRD-03":
        return 25
    # Win theme development — benefits from SAL-01/02/03 intelligence already compiled
    if activity_code == "SAL-04":
        return 20
    # Evaluation criteria — benefits from ITT analysis already done
    if activity_code == "SAL-05":
        return 20

    # MEDIUM compound (15-20%) — downstream with significant prep benefit
    if activity_code in ("SOL-06", "SOL-07", "SOL-08", "SOL-09"):
        return 15  # Benefit from SOL-03/04 design context
    if activity_code in ("COM-01", "COM-03", "COM-04", "COM-05", "COM-06"):
        return 15  # Benefit from solution + requirements context
    if activity_code in ("DEL-02", "DEL-03", "DEL-04"):
        return 20  # Benefit from SOL-07 transition plan + solution design
    if activity_code == "DEL-01":
        return 15  # Risk register benefits from upstream risk flags
    if activity_code in ("PRD-01", "PRD-04"):
        return 20  # Compliance mapping + pricing schedules benefit from upstream
    if activity_code in ("BM-14", "BM-15"):
        return 15  # Clarification management benefits from requirements baseline
    if activity_code == "BM-16":
        return 15  # Win strategy refresh benefits from full intelligence context

    # Review tasks — benefit from higher quality upstream content
    if activity_code in ("PRD-05", "PRD-06", "PRD-07"):
        return 15  # Better first drafts = faster reviews = less rework
    if activity_code in ("PRD-08", "PRD-09"):
        return 10  # QA/submission — modest benefit

    # LOW compound (5-10%)
    if ws in ("LEG",):
        return 10  # Legal benefits from contract analysis context
    if ws in ("SUP",):
        return 10  # Supply chain benefits from solution context
    if activity_code in ("SOL-10", "SOL-11", "SOL-12"):
        return 10  # Evidence/lock/risk — modest upstream benefit
    if activity_code in ("COM-07",):
        return 10

    # Governance — meeting time barely compresses but prep is faster
    if ws == "GOV":
        return 10

    # Post-submission
    if ws == "POST":
        if activity_code in ("POST-01", "POST-04", "POST-05"):
            return 15  # Presentation + clarifications benefit from full bid context
        return 5

    # BM general
    if ws == "BM":
        return 10

    # SAL remaining
    if ws == "SAL":
        if activity_code in ("SAL-02", "SAL-03"):
            return 10  # Benefit from SAL-01 intelligence
        if activity_code == "SAL-07":
            return 15  # Clarification strategy benefits from full requirements knowledge
        if activity_code == "SAL-10":
            return 10
        return 5

    return 5  # Default minimal compound effect


# ── Inference functions for tasks without explicit assessment ────────────────

def infer_what_ai_does(task_name, rating):
    name_lower = task_name.lower()
    if rating == "High":
        if "consolidat" in name_lower or "assemble" in name_lower:
            return "AI consolidates upstream outputs into structured document automatically"
        if "extract" in name_lower or "decompose" in name_lower:
            return "AI extracts and structures data from source documents"
        if "map" in name_lower:
            return "AI maps and cross-references data systematically"
        if "track" in name_lower:
            return "AI automates tracking from structured data sources"
        if "format" in name_lower or "proof" in name_lower:
            return "AI automates formatting, checking, and quality assurance"
        if "harvest" in name_lower:
            return "AI harvests and consolidates data from upstream outputs"
        return "AI executes task autonomously from structured inputs. Human reviews."
    if rating == "Medium":
        if "draft" in name_lower or "produce" in name_lower or "develop" in name_lower:
            return "AI produces first draft. Human refines with domain judgement."
        if "design" in name_lower:
            return "AI proposes options based on requirements. Human makes design decisions."
        if "assess" in name_lower or "analys" in name_lower or "review" in name_lower:
            return "AI provides analysis and assessment. Human validates with judgement."
        if "model" in name_lower:
            return "AI runs models from structured inputs. Human validates assumptions."
        if "plan" in name_lower:
            return "AI drafts plan from upstream outputs. Human validates and refines."
        return "AI accelerates with draft/analysis. Human applies domain judgement."
    # Low
    if "negotiate" in name_lower:
        return "Cannot be AI'd. Negotiation is fundamentally human."
    if "present" in name_lower or "deliver" in name_lower:
        return "AI prepares materials. Delivery is human."
    if "validate" in name_lower or "confirm" in name_lower or "approve" in name_lower:
        return "AI prepares pack/summary. Decision is human governance."
    if "rehearsal" in name_lower:
        return "AI can simulate Q&A. Practice is human."
    return "AI provides minimal preparation support. Task is fundamentally human."


def infer_inputs(task_id, activity_code):
    return f"Upstream {activity_code} outputs"


def infer_value(rating):
    if rating == "High":
        return "Time reduction: Task largely automated from structured data"
    if rating == "Medium":
        return "Decision velocity: AI draft available immediately for human refinement"
    return "Minimal: Preparation support only"


def infer_time_reduction(rating):
    if rating == "High":
        return 60
    if rating == "Medium":
        return 35
    return 10


def infer_decision_velocity(rating, task_name):
    name_lower = task_name.lower()
    if rating == "High" and any(kw in name_lower for kw in ["extract", "analys", "map", "categorise"]):
        return "High"
    if rating == "High":
        return "Medium"
    if rating == "Medium":
        return "Medium"
    if "approve" in name_lower or "confirm" in name_lower:
        return "None"
    return "Low"


def infer_quality_uplift(rating, task_name):
    name_lower = task_name.lower()
    if rating == "High" and any(kw in name_lower for kw in ["cross-reference", "map", "extract", "analys"]):
        return "Yes"
    if rating == "High":
        return "Moderate"
    if rating == "Medium":
        return "Moderate"
    return "Low"


def infer_cost_reduction(rating):
    if rating == "High":
        return "Yes"
    if rating == "Medium":
        return "Moderate"
    return "Low"


def infer_blocker(rating, task_name):
    name_lower = task_name.lower()
    if rating == "Low":
        if "negotiate" in name_lower:
            return "Negotiation is fundamentally human"
        if "present" in name_lower or "deliver" in name_lower:
            return "Client-facing interaction is human"
        if "approve" in name_lower or "confirm" in name_lower or "lock" in name_lower:
            return "Governance decision is human"
        if "validate" in name_lower:
            return "Team validation requires human judgement"
        return "Human relationship/judgement required"
    if rating == "Medium":
        return "Domain judgement required for validation"
    return "Structured data must be available"


def infer_external_data(task_id, task_name):
    name_lower = task_name.lower()
    if "contracts finder" in name_lower or ("procurement" in name_lower and "gather" in name_lower):
        return "Contracts Finder / FTS API"
    if "companies house" in name_lower:
        return "Companies House API"
    if "itt" in name_lower or ("requirement" in name_lower and "extract" in name_lower):
        return "ITT Document Pack"
    if "evidence" in name_lower or "credential" in name_lower:
        return "Evidence Library"
    if "cost model" in name_lower or "price" in name_lower:
        return "Corporate Cost Data"
    if "client" in name_lower and ("strategic" in name_lower or "intelligence" in name_lower):
        return "Client Published Documents"
    if "market" in name_lower or "sector" in name_lower:
        return "Market Intelligence Sources"
    if "competitor" in name_lower:
        return "Companies House API"
    return "None"


# ── Derive value dimensions from detailed assessment text ───────────────────

def derive_decision_velocity(value_text, rating):
    if not value_text:
        return "Medium" if rating != "Low" else "Low"
    vl = value_text.lower()
    if "decision velocity" in vl:
        return "High"
    if rating == "High":
        return "High"
    if rating == "Medium":
        return "Medium"
    return "Low"


def derive_quality_uplift(value_text, rating):
    if not value_text:
        return "Moderate" if rating != "Low" else "Low"
    vl = value_text.lower()
    if "quality uplift" in vl:
        return "Yes"
    if rating == "High":
        return "Moderate"
    return "Low"


def derive_cost_reduction(rating):
    if rating == "High":
        return "Yes"
    if rating == "Medium":
        return "Moderate"
    return "Low"


def derive_blocker_from_value(value_text):
    if not value_text:
        return ""
    vl = value_text.lower()
    if "blocker:" in vl:
        idx = vl.find("blocker:")
        return value_text[idx + 8:].strip()
    return ""


# ── Assign ratings to tasks using activity-level summary distribution ───────

def assign_ratings_from_summary(activity_code, task_ids):
    """Given an activity with known H/M/L counts, assign ratings to tasks."""
    summary = ACTIVITY_SUMMARIES.get(activity_code)
    if not summary:
        return {}
    high, med, low, _, _ = summary
    total = high + med + low
    n_tasks = len(task_ids)

    ratings = {}
    if n_tasks == total:
        # Exact match: assign H first, then M, then L (by task order)
        for tid in task_ids[:high]:
            ratings[tid] = "High"
        for tid in task_ids[high:high + med]:
            ratings[tid] = "Medium"
        for tid in task_ids[high + med:]:
            ratings[tid] = "Low"
    else:
        # Different count: distribute proportionally
        for i, tid in enumerate(task_ids):
            ratio = i / max(n_tasks - 1, 1)
            if ratio < high / max(total, 1):
                ratings[tid] = "High"
            elif ratio < (high + med) / max(total, 1):
                ratings[tid] = "Medium"
            else:
                ratings[tid] = "Low"
    return ratings


# ── Build full task data ────────────────────────────────────────────────────

WS_ORDER = ["SAL", "SOL", "COM", "LEG", "DEL", "SUP", "BM", "PRD", "GOV", "POST"]

WS_NAMES = {
    "SAL": "Sales & Capture", "SOL": "Solution Design", "COM": "Commercial & Pricing",
    "LEG": "Legal & Contractual", "DEL": "Programme & Delivery", "SUP": "Supply Chain & Partners",
    "BM": "Bid Management", "PRD": "Proposal Production", "GOV": "Internal Governance",
    "POST": "Post-Submission",
}


def build_all_tasks():
    tasks = extract_tasks()
    detailed = extract_detailed_assessments()

    # Group tasks by activity for summary-based distribution
    activity_tasks = {}
    for tid, name in tasks:
        activity = tid.split(".")[0]
        activity_tasks.setdefault(activity, []).append(tid)

    # Pre-compute summary-based ratings
    summary_ratings = {}
    for act_code, tids in activity_tasks.items():
        if act_code in ACTIVITY_SUMMARIES:
            summary_ratings.update(assign_ratings_from_summary(act_code, tids))

    rows = []
    for tid, name in tasks:
        ws = tid.split("-")[0]
        activity = tid.split(".")[0]

        if tid in detailed:
            d = detailed[tid]
            rating = d["rating"]
            what_ai = d["what_ai"]
            inputs_needed = d["inputs"]
            value = d["value"]
            role = d["role"]
            time_red = d["time_reduction"]
            dv = derive_decision_velocity(d.get("value", ""), rating)
            qu = derive_quality_uplift(d.get("value", ""), rating)
            cr = derive_cost_reduction(rating)
            blocker_text = derive_blocker_from_value(d.get("value", ""))
            if not blocker_text:
                blocker_text = infer_blocker(rating, name)
            ext_data = infer_external_data(tid, name)
        else:
            # Use summary-based rating if available, else infer from task name
            if tid in summary_ratings:
                rating = summary_ratings[tid]
            else:
                # Fallback inference from task name
                name_lower = name.lower()
                high_kw = ["extract", "decompose", "scrape", "scan", "compile", "consolidat",
                           "map every", "map requirement", "compliance matrix", "cross-reference",
                           "harvest", "automat", "configure repository", "track bid investment",
                           "assemble", "populate pricing", "format", "proof-read", "spell-check",
                           "compliance check"]
                low_kw = ["validate", "confirm", "approve", "sign-off", "lock", "present",
                          "deliver bid", "negotiate", "rehearsal", "debrief", "workshop",
                          "execute engagement", "dialogue", "governance"]
                if any(kw in name_lower for kw in high_kw):
                    rating = "High"
                elif any(kw in name_lower for kw in low_kw):
                    rating = "Low"
                else:
                    rating = "Medium"

            what_ai = infer_what_ai_does(name, rating)
            inputs_needed = infer_inputs(tid, activity)
            value = infer_value(rating)
            role = ACTIVITY_ROLE_OVERRIDES.get(activity, ROLE_MAP.get(ws, "Bid Manager"))
            time_red = infer_time_reduction(rating)
            dv = infer_decision_velocity(rating, name)
            qu = infer_quality_uplift(rating, name)
            cr = infer_cost_reduction(rating)
            blocker_text = infer_blocker(rating, name)
            ext_data = infer_external_data(tid, name)

        phase = get_phase(tid)

        # Compute effort share and days saved
        act_effort = ACTIVITY_EFFORT.get(activity, {})
        activity_effort_days = act_effort.get("effortDays", 0)
        activity_team_size = act_effort.get("teamSize", 1)
        n_tasks_in_activity = len(activity_tasks.get(activity, [tid]))
        task_effort_share = round(activity_effort_days / max(n_tasks_in_activity, 1), 2)
        compound_accel = get_compound_acceleration(tid, activity, name)
        effective_reduction = 1 - (1 - time_red / 100) * (1 - compound_accel / 100)
        ai_days_saved = round(task_effort_share * effective_reduction, 2)

        rows.append({
            "workstream": ws,
            "activity": activity,
            "task_id": tid,
            "rating": rating,
            "name": name,
            "what_ai": what_ai,
            "inputs": inputs_needed,
            "value": value,
            "role": role,
            "time_reduction": time_red,
            "compound_accel": compound_accel,
            "effective_reduction": round(effective_reduction * 100, 1),
            "decision_velocity": dv,
            "quality_uplift": qu,
            "cost_reduction": cr,
            "blocker": blocker_text,
            "external_data": ext_data,
            "phase": phase,
            "activity_effort_days": activity_effort_days,
            "activity_team_size": activity_team_size,
            "task_effort_share": task_effort_share,
            "ai_days_saved": ai_days_saved,
        })

    return rows


# ── Sheet 1: All Tasks ─────────────────────────────────────────────────────

COMPOUND_HIGHLIGHT_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
EDITABLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def write_all_tasks(wb, rows):
    ws = wb.active
    ws.title = "All Tasks"
    # Columns:
    # A=Workstream, B=Activity Code, C=Task ID, D=AI Rating, E=Task Name,
    # F=What AI Does, G=Inputs, H=Value Prop, I=Role,
    # J=Activity Effort, K=Team Size, L=Task Effort Share (FORMULA),
    # M=Direct AI Reduction % (EDITABLE), N=Compound Acceleration % (EDITABLE),
    # O=Effective Reduction % (FORMULA), P=AI Days Saved (FORMULA),
    # Q=Decision Velocity, R=Quality Uplift, S=Cost Reduction,
    # T=Key Blocker, U=External Data, V=Phase, W=User Notes
    headers = [
        "Workstream", "Activity Code", "Task ID", "AI Rating", "Task Name",
        "What AI Does", "Inputs Needed", "Value Proposition", "Primary Role Impact",
        "Activity Effort (days)", "Team Size",
        "Task Effort Share (days)",       # L — FORMULA
        "Direct AI Reduction %",          # M — EDITABLE (yellow)
        "Compound Acceleration %",        # N — EDITABLE (blue highlight)
        "Effective Reduction %",           # O — FORMULA (green)
        "AI Days Saved",                   # P — FORMULA (green)
        "Decision Velocity Impact", "Quality Uplift",
        "Cost Reduction", "Key Blocker", "External Data Required",
        "Implementation Phase", "User Notes / Challenge",
    ]
    ws.append(headers)

    for i, r in enumerate(rows):
        row_num = i + 2  # Excel row (1-indexed, header is row 1)

        # Static values
        ws.cell(row=row_num, column=1, value=r["workstream"])
        ws.cell(row=row_num, column=2, value=r["activity"])
        ws.cell(row=row_num, column=3, value=r["task_id"])
        ws.cell(row=row_num, column=4, value=r["rating"])
        ws.cell(row=row_num, column=5, value=r["name"])
        ws.cell(row=row_num, column=6, value=r["what_ai"])
        ws.cell(row=row_num, column=7, value=r["inputs"])
        ws.cell(row=row_num, column=8, value=r["value"])
        ws.cell(row=row_num, column=9, value=r["role"])
        ws.cell(row=row_num, column=10, value=r["activity_effort_days"])  # J
        ws.cell(row=row_num, column=11, value=r["activity_team_size"])    # K

        # L = Task Effort Share — FORMULA: Activity Effort / count of tasks in this activity
        ws.cell(row=row_num, column=12).value = f'=J{row_num}/COUNTIF($B:$B,B{row_num})'

        # M = Direct AI Reduction % — EDITABLE (value as decimal for formulas)
        ws.cell(row=row_num, column=13, value=r["time_reduction"] / 100)
        ws.cell(row=row_num, column=13).number_format = '0%'
        ws.cell(row=row_num, column=13).fill = EDITABLE_FILL

        # N = Compound Acceleration % — EDITABLE (value as decimal)
        ws.cell(row=row_num, column=14, value=r["compound_accel"] / 100)
        ws.cell(row=row_num, column=14).number_format = '0%'
        ws.cell(row=row_num, column=14).fill = EDITABLE_FILL
        # Highlight high compound values
        if r["compound_accel"] >= 20:
            ws.cell(row=row_num, column=14).fill = COMPOUND_HIGHLIGHT_FILL

        # O = Effective Reduction % — FORMULA: 1 - (1 - Direct) * (1 - Compound)
        ws.cell(row=row_num, column=15).value = f'=1-(1-M{row_num})*(1-N{row_num})'
        ws.cell(row=row_num, column=15).number_format = '0.0%'
        ws.cell(row=row_num, column=15).fill = FORMULA_FILL

        # P = AI Days Saved — FORMULA: Task Effort Share * Effective Reduction
        ws.cell(row=row_num, column=16).value = f'=L{row_num}*O{row_num}'
        ws.cell(row=row_num, column=16).number_format = '0.00'
        ws.cell(row=row_num, column=16).fill = FORMULA_FILL

        # Remaining static columns
        ws.cell(row=row_num, column=17, value=r["decision_velocity"])
        ws.cell(row=row_num, column=18, value=r["quality_uplift"])
        ws.cell(row=row_num, column=19, value=r["cost_reduction"])
        ws.cell(row=row_num, column=20, value=r["blocker"])
        ws.cell(row=row_num, column=21, value=r["external_data"])
        ws.cell(row=row_num, column=22, value=r["phase"])
        ws.cell(row=row_num, column=23, value="")

    n_headers = len(headers)
    last_row = len(rows) + 1
    style_header(ws, n_headers)
    style_body(ws, last_row, n_headers, rating_col=4)

    # Re-apply editable/formula fills (style_body overwrites)
    for i in range(2, last_row + 1):
        ws.cell(row=i, column=12).fill = FORMULA_FILL    # L
        ws.cell(row=i, column=13).fill = EDITABLE_FILL   # M
        ca_val = ws.cell(row=i, column=14).value
        if isinstance(ca_val, (int, float)) and ca_val >= 0.20:
            ws.cell(row=i, column=14).fill = COMPOUND_HIGHLIGHT_FILL
        else:
            ws.cell(row=i, column=14).fill = EDITABLE_FILL
        ws.cell(row=i, column=15).fill = FORMULA_FILL    # O
        ws.cell(row=i, column=16).fill = FORMULA_FILL    # P

    set_col_widths(ws, [10, 10, 14, 10, 50, 45, 30, 40, 16, 12, 10, 14,
                        12, 14, 14, 12, 16, 12, 12, 30, 22, 14, 25])


# ── Sheet 2: Summary by Workstream ─────────────────────────────────────────

def write_summary(wb, rows):
    ws = wb.create_sheet("Summary by Workstream")
    # All Tasks sheet: A=Workstream, D=Rating, L=Task Effort Share, O=Effective Reduction, P=AI Days Saved
    at = "'All Tasks'"
    headers = [
        "Workstream", "Full Name", "Total Tasks", "High", "Medium",
        "Low", "Methodology Effort (days)", "Avg Effective Reduction %",
        "AI Days Saved", "Effort Reduction %",
    ]
    ws.append(headers)

    for i, wscode in enumerate(WS_ORDER):
        r = i + 2  # Excel row

        ws.cell(row=r, column=1, value=wscode)
        ws.cell(row=r, column=2, value=WS_NAMES.get(wscode, wscode))

        # C = Total tasks — COUNTIF on All Tasks workstream column
        ws.cell(row=r, column=3).value = f'=COUNTIF({at}!$A:$A,A{r})'

        # D = High count
        ws.cell(row=r, column=4).value = f'=COUNTIFS({at}!$A:$A,A{r},{at}!$D:$D,"High")'

        # E = Medium count
        ws.cell(row=r, column=5).value = f'=COUNTIFS({at}!$A:$A,A{r},{at}!$D:$D,"Medium")'

        # F = Low count
        ws.cell(row=r, column=6).value = f'=COUNTIFS({at}!$A:$A,A{r},{at}!$D:$D,"Low")'

        # G = Methodology effort (static — sum of unique activity efforts)
        ws_activities = set(r2["activity"] for r2 in rows if r2["workstream"] == wscode)
        ws_effort = sum(ACTIVITY_EFFORT.get(a, {}).get("effortDays", 0) for a in ws_activities)
        ws.cell(row=r, column=7, value=ws_effort)

        # H = Avg Effective Reduction — AVERAGEIF on All Tasks
        ws.cell(row=r, column=8).value = f'=AVERAGEIF({at}!$A:$A,A{r},{at}!$O:$O)'
        ws.cell(row=r, column=8).number_format = '0.0%'

        # I = AI Days Saved — SUMIF on All Tasks col P
        ws.cell(row=r, column=9).value = f'=SUMIF({at}!$A:$A,A{r},{at}!$P:$P)'
        ws.cell(row=r, column=9).number_format = '0.0'
        ws.cell(row=r, column=9).fill = FORMULA_FILL

        # J = Effort Reduction % — Days Saved / Methodology Effort
        ws.cell(row=r, column=10).value = f'=IF(G{r}>0,I{r}/G{r},0)'
        ws.cell(row=r, column=10).number_format = '0.0%'
        ws.cell(row=r, column=10).fill = FORMULA_FILL

    # Totals row
    tr = len(WS_ORDER) + 2
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=2, value="")
    ws.cell(row=tr, column=3).value = f'=SUM(C2:C{tr-1})'
    ws.cell(row=tr, column=4).value = f'=SUM(D2:D{tr-1})'
    ws.cell(row=tr, column=5).value = f'=SUM(E2:E{tr-1})'
    ws.cell(row=tr, column=6).value = f'=SUM(F2:F{tr-1})'
    ws.cell(row=tr, column=7).value = f'=SUM(G2:G{tr-1})'
    ws.cell(row=tr, column=8).value = f'=AVERAGE({at}!$O:$O)'
    ws.cell(row=tr, column=8).number_format = '0.0%'
    ws.cell(row=tr, column=9).value = f'=SUM(I2:I{tr-1})'
    ws.cell(row=tr, column=9).number_format = '0.0'
    ws.cell(row=tr, column=9).fill = FORMULA_FILL
    ws.cell(row=tr, column=10).value = f'=IF(G{tr}>0,I{tr}/G{tr},0)'
    ws.cell(row=tr, column=10).number_format = '0.0%'
    ws.cell(row=tr, column=10).fill = FORMULA_FILL

    nrows = tr
    style_header(ws, len(headers))
    style_body(ws, nrows, len(headers))
    for col in range(1, len(headers) + 1):
        ws.cell(row=tr, column=col).font = BOLD_FONT
    set_col_widths(ws, [12, 22, 10, 8, 10, 8, 16, 18, 14, 14])


# ── Sheet 3: Top 10 AI Applications ────────────────────────────────────────

TOP_10 = [
    {
        "rank": 1,
        "tasks": "SOL-01.1.1, SOL-01.1.2, SOL-01.1.3",
        "application": "ITT requirements extraction and analysis",
        "value": "Transforms days of manual extraction into hours. Cross-references across documents at scale. Catches hidden requirements.",
        "time_saving": "60-70%",
        "role": "Solution Architect",
        "phase": "Phase 1",
    },
    {
        "rank": 2,
        "tasks": "PRD-02.1.1",
        "application": "AI-assisted response drafting",
        "value": "AI generates first drafts from storyboard + solution design + win themes. Writers refine, not write from scratch.",
        "time_saving": "40-50%",
        "role": "Writers",
        "phase": "Phase 2",
    },
    {
        "rank": 3,
        "tasks": "SAL-05.1.1, SAL-05.1.2, SAL-05.1.3",
        "application": "Evaluation criteria extraction and scoring analysis",
        "value": "Automatically extracts criteria, weightings, grade descriptors. Produces marks concentration analysis.",
        "time_saving": "60-80%",
        "role": "Bid Manager",
        "phase": "Phase 1",
    },
    {
        "rank": 4,
        "tasks": "PRD-01.1.1",
        "application": "Automated compliance mapping",
        "value": "Maps every ITT requirement to response sections. Identifies compliance gaps systematically.",
        "time_saving": "60-70%",
        "role": "Bid Manager",
        "phase": "Phase 1",
    },
    {
        "rank": 5,
        "tasks": "SAL-06.1.2",
        "application": "Rapid ITT documentation analysis",
        "value": "Analyses requirements, contract, payment mechanism, SLAs across the entire ITT pack for go/no-go.",
        "time_saving": "70%",
        "role": "Bid Manager",
        "phase": "Phase 1",
    },
    {
        "rank": 6,
        "tasks": "COM-05.2.1, COM-05.2.2",
        "application": "Automated sensitivity and stress testing",
        "value": "Runs all financial scenarios, tornado analysis, break-even modelling from structured cost data.",
        "time_saving": "70-80%",
        "role": "Commercial Lead",
        "phase": "Phase 4",
    },
    {
        "rank": 7,
        "tasks": "LEG-01.1.1",
        "application": "Contract clause analysis",
        "value": "AI analyses every clause, identifies non-standard terms, compares to corporate playbook, flags risk areas.",
        "time_saving": "50-60%",
        "role": "Legal Lead",
        "phase": "Phase 1",
    },
    {
        "rank": 8,
        "tasks": "SAL-01.1.1, SAL-01.1.2, SAL-01.1.3, SAL-01.1.4",
        "application": "Customer and market intelligence gathering",
        "value": "Automated Contracts Finder scraping, client strategy analysis, supplier landscape mapping.",
        "time_saving": "60-70%",
        "role": "Capture Lead",
        "phase": "Phase 3",
    },
    {
        "rank": 9,
        "tasks": "BM-13.1.1, SOL-12.1.1, COM-07.1.1",
        "application": "Automated risk consolidation",
        "value": "Consolidates all workstream risks, removes duplicates, identifies cross-cutting risks, flags gaps.",
        "time_saving": "50-60%",
        "role": "Bid Manager",
        "phase": "Phase 5",
    },
    {
        "rank": 10,
        "tasks": "SOL-10.1.1, SOL-10.1.2, SOL-10.1.3",
        "application": "Evidence gap analysis and library search",
        "value": "Maps evidence needs to sections, searches library, identifies gaps, prioritises by scoring impact.",
        "time_saving": "60-70%",
        "role": "Bid Manager",
        "phase": "Phase 3",
    },
]


def write_top_10(wb):
    ws = wb.create_sheet("Top 10 AI Applications")
    headers = [
        "Rank", "Task(s)", "AI Application", "Value", "Time Saving",
        "Affected Role", "Implementation Phase", "User Notes",
    ]
    ws.append(headers)
    for item in TOP_10:
        ws.append([
            item["rank"], item["tasks"], item["application"], item["value"],
            item["time_saving"], item["role"], item["phase"], "",
        ])
    style_header(ws, len(headers))
    style_body(ws, 11, len(headers))
    set_col_widths(ws, [6, 35, 35, 55, 12, 18, 16, 25])


# ── Sheet 4: Implementation Roadmap ────────────────────────────────────────

ROADMAP = [
    {
        "phase": "Phase 1: Document Intelligence",
        "skills": "ITT extraction, requirements analysis, contract clause analysis, evaluation criteria extraction",
        "tasks_enabled": "SOL-01, SAL-05, SAL-06.1.2, LEG-01, PRD-01, SAL-07.1.1, SAL-07.2.1",
        "impact": "Highest impact. Transforms the first 2 weeks of every bid. Enables faster go/no-go and more informed strategy.",
    },
    {
        "phase": "Phase 2: Content Generation",
        "skills": "Response drafting, storyboard generation, governance pack assembly, compliance mapping",
        "tasks_enabled": "PRD-02, BM-10, GOV-01-06, PRD-01, BM-06, PRD-08, SAL-04.2.2, SAL-06.2.4, SAL-06.4.1",
        "impact": "Highest visibility. AI writes first drafts. Bid team sees immediate productivity gain.",
    },
    {
        "phase": "Phase 3: Intelligence & Research",
        "skills": "Market scanning, competitor profiling, client intelligence, evidence library search",
        "tasks_enabled": "SAL-01, SAL-02, SAL-03, SAL-10, SOL-02, SOL-10",
        "impact": "Strategic value. Better intelligence, faster capture planning.",
    },
    {
        "phase": "Phase 4: Financial Modelling",
        "skills": "Cost modelling, sensitivity analysis, price-to-win, margin modelling",
        "tasks_enabled": "COM-01, COM-02, COM-05, SAL-02.2.1, SAL-02.3.4",
        "impact": "Commercial value. Automated financial modelling with scenario testing.",
    },
    {
        "phase": "Phase 5: Risk & Consolidation",
        "skills": "Risk register consolidation, assumption tracking, clarification impact analysis",
        "tasks_enabled": "BM-13, SOL-12, COM-07, DEL-06, BM-15",
        "impact": "Quality uplift. No risks fall through cracks.",
    },
]


def write_roadmap(wb):
    ws = wb.create_sheet("Implementation Roadmap")
    headers = ["Phase", "Skills to Build", "Tasks Enabled", "Business Impact", "User Notes"]
    ws.append(headers)
    for item in ROADMAP:
        ws.append([item["phase"], item["skills"], item["tasks_enabled"], item["impact"], ""])
    style_header(ws, len(headers))
    style_body(ws, 6, len(headers))
    set_col_widths(ws, [28, 55, 55, 60, 25])


# ── Sheet 5: External Data Sources ─────────────────────────────────────────

EXTERNAL_DATA = [
    {
        "source": "Contracts Finder / FTS API",
        "applications": "SAL-01 (procurement intel), SAL-02 (incumbent pricing), SAL-03 (competitor identification), COM-02 (price benchmarks)",
        "priority": "Critical",
        "availability": "Public API, free access. FTS for EU/international procurements.",
    },
    {
        "source": "Companies House API",
        "applications": "SAL-03 (competitor profiling), SUP-01 (partner due diligence)",
        "priority": "High",
        "availability": "Public API, free access.",
    },
    {
        "source": "ITT Document Pack (uploaded)",
        "applications": "SOL-01 (requirements), SAL-05 (evaluation criteria), SAL-06 (ITT analysis), LEG-01 (contract review), PRD-01 (compliance)",
        "priority": "Critical",
        "availability": "Client-provided at ITT receipt. Must be uploaded to PWIN.",
    },
    {
        "source": "Client Published Documents",
        "applications": "SAL-01 (strategic context), SAL-02 (performance data)",
        "priority": "High",
        "availability": "Public domain. Requires web scraping or manual collection.",
    },
    {
        "source": "Evidence Library (internal)",
        "applications": "SAL-04 (theme substantiation), SOL-10 (evidence strategy), PRD-03 (evidence assembly)",
        "priority": "Critical",
        "availability": "Internal. Must be structured and accessible to AI.",
    },
    {
        "source": "Corporate Cost Data",
        "applications": "COM-01 (rate cards, overhead rates), COM-02 (benchmarks)",
        "priority": "High",
        "availability": "Internal. Finance team provides rate cards and overhead model.",
    },
    {
        "source": "Industry Benchmarks (Gartner, ISG)",
        "applications": "COM-02 (market rates), SAL-02 (performance benchmarks)",
        "priority": "Medium",
        "availability": "Commercial subscription required. Analyst reports available.",
    },
    {
        "source": "CRM / Relationship Data",
        "applications": "SAL-01 (client intel), SAL-10 (stakeholder mapping)",
        "priority": "Medium",
        "availability": "Internal CRM system. Requires integration or export.",
    },
]


def write_external_data(wb):
    ws = wb.create_sheet("External Data Sources")
    headers = ["Data Source", "AI Applications Enabled", "Priority", "Availability", "User Notes"]
    ws.append(headers)
    for item in EXTERNAL_DATA:
        ws.append([item["source"], item["applications"], item["priority"], item["availability"], ""])
    style_header(ws, len(headers))
    style_body(ws, len(EXTERNAL_DATA) + 1, len(headers))
    set_col_widths(ws, [30, 65, 12, 55, 25])


# ── Sheet 6: Effort Reduction Summary (by Activity) ─────────────────────────

GOLD_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="253551", end_color="253551", fill_type="solid")
SUBTOTAL_FONT = Font(name="Calibri", size=9, bold=True, color="C49438")
TOTAL_FILL = PatternFill(start_color="C49438", end_color="C49438", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color="1B2A4A")
PCT_HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PCT_MED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PCT_LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def write_effort_reduction(wb, rows):
    ws = wb.create_sheet("Effort Reduction Summary")
    at = "'All Tasks'"

    headers = [
        "Workstream", "Activity Code", "Activity Name",
        "Effort (person-days)", "Team Size", "L3 Tasks",
        "Avg Effective Reduction %", "AI Days Saved",
        "Remaining Effort (days)", "Effort Reduction %",
        "Methodology Note",
    ]
    ws.append(headers)

    current_row = 2
    grand_total_effort = 0
    activity_rows = []  # Track row numbers for subtotals

    for wscode in WS_ORDER:
        ws_rows = [r for r in rows if r["workstream"] == wscode]
        if not ws_rows:
            continue

        seen = set()
        activities = []
        for r in ws_rows:
            if r["activity"] not in seen:
                seen.add(r["activity"])
                activities.append(r["activity"])

        ws_start_row = current_row
        ws_effort_total = 0

        for act_code in activities:
            act_data = ACTIVITY_EFFORT.get(act_code, {})
            effort = act_data.get("effortDays", 0)
            team = act_data.get("teamSize", 1)
            act_name = act_data.get("name", act_code)

            note = ""
            if effort == 0:
                note = "Continuous/tracking — effort embedded in bid management overhead"

            r = current_row
            ws.cell(row=r, column=1, value=wscode)
            ws.cell(row=r, column=2, value=act_code)
            ws.cell(row=r, column=3, value=act_name)
            ws.cell(row=r, column=4, value=effort)
            ws.cell(row=r, column=5, value=team)

            # F = L3 Tasks count — COUNTIF on All Tasks activity column
            ws.cell(row=r, column=6).value = f'=COUNTIF({at}!$B:$B,B{r})'

            # G = Avg Effective Reduction — AVERAGEIF on All Tasks
            ws.cell(row=r, column=7).value = f'=AVERAGEIF({at}!$B:$B,B{r},{at}!$O:$O)'
            ws.cell(row=r, column=7).number_format = '0.0%'

            # H = AI Days Saved — SUMIF on All Tasks col P
            ws.cell(row=r, column=8).value = f'=SUMIF({at}!$B:$B,B{r},{at}!$P:$P)'
            ws.cell(row=r, column=8).number_format = '0.0'
            ws.cell(row=r, column=8).fill = FORMULA_FILL

            # I = Remaining Effort — Effort - Days Saved
            ws.cell(row=r, column=9).value = f'=D{r}-H{r}'
            ws.cell(row=r, column=9).number_format = '0.0'
            ws.cell(row=r, column=9).fill = FORMULA_FILL

            # J = Effort Reduction %
            ws.cell(row=r, column=10).value = f'=IF(D{r}>0,H{r}/D{r},0)'
            ws.cell(row=r, column=10).number_format = '0.0%'
            ws.cell(row=r, column=10).fill = FORMULA_FILL

            ws.cell(row=r, column=11, value=note)

            activity_rows.append(current_row)
            current_row += 1
            ws_effort_total += effort
            grand_total_effort += effort

        # Workstream subtotal row
        r = current_row
        ws.cell(row=r, column=1, value="")
        ws.cell(row=r, column=2, value=f"{wscode} SUBTOTAL")
        ws.cell(row=r, column=3, value=WS_NAMES.get(wscode, ""))
        ws.cell(row=r, column=4).value = f'=SUM(D{ws_start_row}:D{r-1})'
        ws.cell(row=r, column=5, value="")
        ws.cell(row=r, column=6, value="")
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8).value = f'=SUM(H{ws_start_row}:H{r-1})'
        ws.cell(row=r, column=8).number_format = '0.0'
        ws.cell(row=r, column=9).value = f'=D{r}-H{r}'
        ws.cell(row=r, column=9).number_format = '0.0'
        ws.cell(row=r, column=10).value = f'=IF(D{r}>0,H{r}/D{r},0)'
        ws.cell(row=r, column=10).number_format = '0.0%'
        ws.cell(row=r, column=11, value="")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = SUBTOTAL_FILL
            cell.font = SUBTOTAL_FONT
            cell.border = THIN_BORDER
        current_row += 1

    # Grand total row
    r = current_row
    ws.cell(row=r, column=1, value="")
    ws.cell(row=r, column=2, value="TOTAL")
    ws.cell(row=r, column=3, value="All 84 Activities")
    ws.cell(row=r, column=4, value=grand_total_effort)
    ws.cell(row=r, column=5, value="")
    ws.cell(row=r, column=6, value=len(rows))
    ws.cell(row=r, column=7, value="")

    # Sum all activity-level days saved (sum the SUMIF cells, not a SUMIF itself)
    # Use SUM of all task-level days saved from All Tasks
    ws.cell(row=r, column=8).value = f'=SUM({at}!$P:$P)'
    ws.cell(row=r, column=8).number_format = '0.0'
    ws.cell(row=r, column=9).value = f'=D{r}-H{r}'
    ws.cell(row=r, column=9).number_format = '0.0'
    ws.cell(row=r, column=10).value = f'=IF(D{r}>0,H{r}/D{r},0)'
    ws.cell(row=r, column=10).number_format = '0.0%'
    ws.cell(row=r, column=11, value="")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER

    style_header(ws, len(headers))
    for row in range(2, current_row):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.font or not cell.font.bold:
                cell.font = BODY_FONT
            cell.alignment = WRAP
            if not cell.border or cell.border == Border():
                cell.border = THIN_BORDER

    set_col_widths(ws, [12, 14, 45, 14, 10, 10, 16, 14, 16, 14, 50])

    return grand_total_effort, 0, 0  # Actuals now computed by formulas


# ── Real-world resource profile data ──────────────────────────────────────

RESOURCE_PLAN = {
    "SRIE / Bid Sponsor":              [1,1,1,1,1,2,2,2,2,2,2,2,2,2,2,10,10,10],
    "Bid Director":                    [1,1,1,20,20,20,20,20,20,20,20,20,2,2,2,5,5,5],
    "Bid / Pursuit Lead":              [5,5,5,10,10,10,10,10,10,10,10,10,2,2,2,5,5,5],
    "Bid Manager":                     [0.5,0.5,0.5,20,20,20,20,20,20,20,20,20,2,2,2,0,0,0],
    "Technical / Solution Owner":      [2,2,2,5,5,20,20,20,20,20,20,20,5,5,5,10,10,10],
    "Delivery / Transformation Dir":   [2,2,2,5,5,20,20,20,20,20,20,20,5,5,5,20,20,20],
    "Commercial Lead":                 [2,2,2,5,5,10,10,10,10,10,10,10,5,5,5,2,10,10],
    "Solution Architect":              [0,0,0,5,5,20,20,20,20,20,20,20,0,0,0,5,5,5],
    "SME":                             [1,1,1,10,10,60,60,60,60,60,60,60,2,2,2,5,5,5],
    "Bid Writer / Senior Bid Writer":  [0,0,0,0,0,0,0,0,40,40,40,40,0,0,0,0,0,0],
    "Bid Coordinator":                 [0,0,0,10,10,10,10,10,10,10,10,10,0.5,0.5,0.5,0,0,0],
    "Legal":                           [0,0,0,5,1,1,1,1,1,1,1,1,2,10,2,5,5,5],
}

RESOURCE_DAY_RATES = {
    "SRIE / Bid Sponsor": 1600,
    "Bid Director": 850,
    "Bid / Pursuit Lead": 850,
    "Bid Manager": 650,
    "Technical / Solution Owner": 850,
    "Delivery / Transformation Dir": 850,
    "Commercial Lead": 850,
    "Solution Architect": 850,
    "SME": 850,
    "Bid Writer / Senior Bid Writer": 550,
    "Bid Coordinator": 350,
    "Legal": 650,
}

RESOURCE_PHASES = [
    "Capture","Capture","Capture",
    "Qualification","Qualification",
    "Proposal Development","Proposal Development","Proposal Development",
    "Proposal Development","Proposal Development","Proposal Development",
    "Submission/Review",
    "Clarifications/Standstill",
    "Contract Finalisation","Contract Finalisation",
    "Handover","Handover","Handover",
]

# Role-to-workstream mapping (weighted by where the role's effort primarily sits)
ROLE_TO_WORKSTREAM = {
    "SRIE / Bid Sponsor":              [("GOV", 0.5), ("SAL", 0.3), ("BM", 0.2)],
    "Bid Director":                    [("GOV", 0.3), ("SAL", 0.3), ("BM", 0.4)],
    "Bid / Pursuit Lead":              [("SAL", 0.6), ("BM", 0.4)],
    "Bid Manager":                     [("BM", 0.5), ("PRD", 0.3), ("GOV", 0.1), ("POST", 0.1)],
    "Technical / Solution Owner":      [("SOL", 0.8), ("DEL", 0.2)],
    "Delivery / Transformation Dir":   [("DEL", 0.7), ("SOL", 0.2), ("POST", 0.1)],
    "Commercial Lead":                 [("COM", 0.8), ("SOL", 0.1), ("POST", 0.1)],
    "Solution Architect":              [("SOL", 0.9), ("PRD", 0.1)],
    "SME":                             [("SOL", 0.4), ("PRD", 0.4), ("SUP", 0.2)],
    "Bid Writer / Senior Bid Writer":  [("PRD", 1.0)],
    "Bid Coordinator":                 [("PRD", 0.5), ("BM", 0.3), ("SUP", 0.2)],
    "Legal":                           [("LEG", 0.8), ("COM", 0.1), ("POST", 0.1)],
}


# ── Sheet 7: Real-World AI Impact ────────────────────────────────────────

RW_HEADER_FILL = PatternFill(start_color="0D2137", end_color="0D2137", fill_type="solid")
RW_HEADER_FONT = Font(name="Calibri", bold=True, color="C49438", size=10)
RW_SECTION_FILL = PatternFill(start_color="162032", end_color="162032", fill_type="solid")
RW_SECTION_FONT = Font(name="Calibri", bold=True, color="E8ECF1", size=10)
RW_BODY_FONT = Font(name="Calibri", size=9, color="94A3B8")
RW_NUM_FONT = Font(name="Calibri", size=9, color="E8ECF1")
RW_GOLD_FONT = Font(name="Calibri", size=10, bold=True, color="C49438")
RW_BG = PatternFill(start_color="0F1724", end_color="0F1724", fill_type="solid")
RW_BORDER = Border(
    left=Side(style="thin", color="253551"),
    right=Side(style="thin", color="253551"),
    top=Side(style="thin", color="253551"),
    bottom=Side(style="thin", color="253551"),
)


def style_rw_cell(cell, font=None, fill=None, number_format=None):
    cell.font = font or RW_BODY_FONT
    cell.fill = fill or RW_BG
    cell.alignment = WRAP
    cell.border = RW_BORDER
    if number_format:
        cell.number_format = number_format


def write_real_world_impact(wb, rows):  # noqa: C901 — complex but sequential
    ws = wb.create_sheet("Real-World AI Impact")
    ws.sheet_properties.tabColor = "C49438"

    # ── Section 1: Assumptions ──
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value="REAL-WORLD AI IMPACT — COMPLEX SERVICES BID")
    style_rw_cell(c, font=RW_GOLD_FONT, fill=RW_HEADER_FILL)
    for col in range(2, 11):
        style_rw_cell(ws.cell(row=row, column=col), fill=RW_HEADER_FILL)

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    assumptions = (
        "Assumptions: (1) Five phases: Capture → Qualification → Proposal Development → "
        "Submission/Review → Contract Award/Handover. (2) Full team across all workstreams (12 roles). "
        "(3) Base cost only (excl. 1.4x company uplift). (4) 18-month pursuit timeline. "
        "(5) Resource plan days are authoritative (2,050 person-days). "
        "(6) AI reduction % from bottom-up assessment of 295 L3 methodology tasks."
    )
    c = ws.cell(row=row, column=1, value=assumptions)
    style_rw_cell(c, font=Font(name="Calibri", size=8, italic=True, color="64748B"), fill=RW_BG)
    for col in range(2, 11):
        style_rw_cell(ws.cell(row=row, column=col), fill=RW_BG)
    ws.row_dimensions[row].height = 40

    # ── Section 2: Per-Role Summary ──
    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value="RESOURCE PROFILE — PER ROLE")
    style_rw_cell(c, font=RW_SECTION_FONT, fill=RW_SECTION_FILL)
    for col in range(2, 11):
        style_rw_cell(ws.cell(row=row, column=col), fill=RW_SECTION_FILL)

    row = 5
    role_headers = ["Role", "Base Day Rate", "Total Days", "Base Cost", "% of Total",
                    "Primary Workstream(s)", "", "", "", ""]
    for col, h in enumerate(role_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        style_rw_cell(c, font=RW_HEADER_FONT, fill=RW_HEADER_FILL)

    role_totals = {}
    for role, allocs in RESOURCE_PLAN.items():
        days = sum(allocs)
        rate = RESOURCE_DAY_RATES[role]
        cost = days * rate
        role_totals[role] = {"days": days, "rate": rate, "cost": cost}

    total_rw_days = sum(r["days"] for r in role_totals.values())
    total_rw_cost = sum(r["cost"] for r in role_totals.values())

    row = 6
    for role in RESOURCE_PLAN:
        d = role_totals[role]
        pct = d["days"] / total_rw_days * 100
        ws_list = ", ".join(w for w, _ in ROLE_TO_WORKSTREAM[role])
        vals = [role, d["rate"], d["days"], d["cost"], pct / 100, ws_list, "", "", "", ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            if col == 2:
                style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
            elif col == 4:
                style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
            elif col == 5:
                style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0%")
            elif col == 3:
                style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0")
            else:
                style_rw_cell(c, font=RW_BODY_FONT)
        row += 1

    # Role totals
    vals = ["TOTAL", "", total_rw_days, total_rw_cost, 1.0, "", "", "", "", ""]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        if col in (3,):
            style_rw_cell(c, font=RW_GOLD_FONT, number_format="0.0")
        elif col == 4:
            style_rw_cell(c, font=RW_GOLD_FONT, number_format="£#,##0")
        elif col == 5:
            style_rw_cell(c, font=RW_GOLD_FONT, number_format="0.0%")
        else:
            style_rw_cell(c, font=RW_GOLD_FONT)
    row += 2

    # ── Section 3: Per-Phase Summary ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value="EFFORT BY PHASE")
    style_rw_cell(c, font=RW_SECTION_FONT, fill=RW_SECTION_FILL)
    for col in range(2, 11):
        style_rw_cell(ws.cell(row=row, column=col), fill=RW_SECTION_FILL)
    row += 1

    phase_headers = ["Phase", "Months", "Days", "Base Cost", "% of Total",
                     "", "", "", "", ""]
    for col, h in enumerate(phase_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        style_rw_cell(c, font=RW_HEADER_FONT, fill=RW_HEADER_FILL)
    row += 1

    phase_order = ["Capture", "Qualification", "Proposal Development",
                   "Submission/Review", "Clarifications/Standstill",
                   "Contract Finalisation", "Handover"]
    for phase in phase_order:
        phase_months = [i for i, p in enumerate(RESOURCE_PHASES) if p == phase]
        n_months = len(phase_months)
        phase_days = 0
        phase_cost = 0
        for role, allocs in RESOURCE_PLAN.items():
            rate = RESOURCE_DAY_RATES[role]
            for m in phase_months:
                phase_days += allocs[m]
                phase_cost += allocs[m] * rate
        pct = phase_days / total_rw_days
        vals = [phase, n_months, phase_days, phase_cost, pct, "", "", "", "", ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            if col == 3:
                style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0")
            elif col == 4:
                style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
            elif col == 5:
                style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0%")
            else:
                style_rw_cell(c, font=RW_BODY_FONT)
        row += 1
    row += 1

    # ── Section 4: Workstream Analysis — Real vs Methodology vs AI ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value="WORKSTREAM ANALYSIS — REAL-WORLD vs METHODOLOGY vs AI SAVING")
    style_rw_cell(c, font=RW_SECTION_FONT, fill=RW_SECTION_FILL)
    for col in range(2, 11):
        style_rw_cell(ws.cell(row=row, column=col), fill=RW_SECTION_FILL)
    row += 1

    ws_headers = ["Workstream", "Real Days", "Real Cost", "Methodology Days",
                  "Multiplier", "AI Reduction %", "Days Saved", "Cost Saved",
                  "Remaining Days", "Remaining Cost"]
    for col, h in enumerate(ws_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        style_rw_cell(c, font=RW_HEADER_FONT, fill=RW_HEADER_FILL)
    row += 1

    # Compute workstream distribution (static real-world allocation)
    ws_days_rw = {w: 0.0 for w in WS_ORDER}
    ws_cost_rw = {w: 0.0 for w in WS_ORDER}
    for role, data in role_totals.items():
        for w, share in ROLE_TO_WORKSTREAM[role]:
            ws_days_rw[w] += data["days"] * share
            ws_cost_rw[w] += data["cost"] * share

    # Methodology effort per workstream
    meth_ws = {}
    for act, v in ACTIVITY_EFFORT.items():
        w = act.split("-")[0]
        meth_ws[w] = meth_ws.get(w, 0) + v["effortDays"]

    # Summary sheet reference for AI reduction %
    summ = "'Summary by Workstream'"
    # Summary sheet: A=workstream code, J=Effort Reduction % (rows 2-11 for SAL-POST)
    # We'll use VLOOKUP or direct cell refs. Rows 2-11 map to WS_ORDER.

    ws_start = row
    for i, w in enumerate(WS_ORDER):
        rd = ws_days_rw[w]
        rc = ws_cost_rw[w]
        md = meth_ws.get(w, 0)

        ws.cell(row=row, column=1, value=WS_NAMES.get(w, w))
        style_rw_cell(ws.cell(row=row, column=1), font=RW_BODY_FONT)

        # B = Real Days (static)
        c = ws.cell(row=row, column=2, value=rd)
        style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0")

        # C = Real Cost (static)
        c = ws.cell(row=row, column=3, value=rc)
        style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")

        # D = Methodology Days (static)
        c = ws.cell(row=row, column=4, value=md)
        style_rw_cell(c, font=RW_NUM_FONT)

        # E = Multiplier (formula)
        c = ws.cell(row=row, column=5)
        c.value = f'=IF(D{row}>0,B{row}/D{row},0)'
        style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0\"x\"")

        # F = AI Reduction % — FORMULA referencing Summary by Workstream
        # Summary row for this workstream = i + 2
        summ_row = i + 2
        c = ws.cell(row=row, column=6)
        c.value = f'={summ}!J{summ_row}'
        style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0%")

        # G = Days Saved (formula: Real Days * AI %)
        c = ws.cell(row=row, column=7)
        c.value = f'=B{row}*F{row}'
        style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0")

        # H = Cost Saved (formula: Real Cost * AI %)
        c = ws.cell(row=row, column=8)
        c.value = f'=C{row}*F{row}'
        style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")

        # I = Remaining Days
        c = ws.cell(row=row, column=9)
        c.value = f'=B{row}-G{row}'
        style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0")

        # J = Remaining Cost
        c = ws.cell(row=row, column=10)
        c.value = f'=C{row}-H{row}'
        style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")

        row += 1

    # Workstream totals row
    ws_end = row - 1
    total_labels = ["TOTAL"]
    c = ws.cell(row=row, column=1, value="TOTAL")
    style_rw_cell(c, font=RW_GOLD_FONT)
    for col, fmt in [(2, "0.0"), (3, "£#,##0"), (4, "0"), (7, "0.0"), (8, "£#,##0"), (9, "0.0"), (10, "£#,##0")]:
        c = ws.cell(row=row, column=col)
        c.value = f'=SUM({get_column_letter(col)}{ws_start}:{get_column_letter(col)}{ws_end})'
        style_rw_cell(c, font=RW_GOLD_FONT, number_format=fmt)

    # E = overall multiplier
    c = ws.cell(row=row, column=5)
    c.value = f'=IF(D{row}>0,B{row}/D{row},0)'
    style_rw_cell(c, font=RW_GOLD_FONT, number_format="0.0\"x\"")

    # F = overall AI %
    c = ws.cell(row=row, column=6)
    c.value = f'=IF(B{row}>0,G{row}/B{row},0)'
    style_rw_cell(c, font=RW_GOLD_FONT, number_format="0.0%")

    row += 2

    # ── Section 5: Four Value Layers ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value="FOUR VALUE LAYERS — HEADLINE NUMBERS (all formula-driven)")
    style_rw_cell(c, font=RW_SECTION_FONT, fill=RW_SECTION_FILL)
    for col in range(2, 11):
        style_rw_cell(ws.cell(row=row, column=col), fill=RW_SECTION_FILL)
    row += 1

    # Capture + Qualification cost (static — pre-gate spend)
    pre_gate_cost = 0
    for role, allocs in RESOURCE_PLAN.items():
        rate = RESOURCE_DAY_RATES[role]
        for m in range(5):
            pre_gate_cost += allocs[m] * rate

    # ws_end+1 is the TOTAL row from the workstream section above
    total_row_ref = ws_end + 1  # The TOTAL row in workstream analysis

    # Header row
    layer_headers = ["LAYER", "Description", "Metric", "Base Value", "Loaded (1.4x)",
                     "Annual (3 bids)", "Annual (5 bids)", "Basis", "", ""]
    for col, h in enumerate(layer_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        style_rw_cell(c, font=RW_HEADER_FONT, fill=RW_HEADER_FILL)
    row += 1

    # Layer 1 — Task Efficiency: Days saved
    r = row
    ws.cell(row=r, column=1, value="Layer 1"); style_rw_cell(ws.cell(row=r, column=1), font=Font(name="Calibri", size=9, bold=True, color="C49438"))
    ws.cell(row=r, column=2, value="Task Efficiency"); style_rw_cell(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value="AI person-days saved"); style_rw_cell(ws.cell(row=r, column=3))
    c = ws.cell(row=r, column=4); c.value = f'=G{total_row_ref}'; style_rw_cell(c, font=RW_NUM_FONT, number_format="0.0")
    for col in range(5, 11): style_rw_cell(ws.cell(row=r, column=col))
    ws.cell(row=r, column=8, value="Formula: references workstream AI days saved total"); style_rw_cell(ws.cell(row=r, column=8))
    row += 1

    # Layer 1 — Cost saving
    r = row
    ws.cell(row=r, column=1, value=""); style_rw_cell(ws.cell(row=r, column=1))
    ws.cell(row=r, column=2, value="  Cost saving"); style_rw_cell(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value="Base cost saved"); style_rw_cell(ws.cell(row=r, column=3))
    c = ws.cell(row=r, column=4); c.value = f'=H{total_row_ref}'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    c = ws.cell(row=r, column=5); c.value = f'=D{r}*1.4'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    c = ws.cell(row=r, column=6); c.value = f'=D{r}*3'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    c = ws.cell(row=r, column=7); c.value = f'=D{r}*5'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    ws.cell(row=r, column=8, value="AI % applied to real-world cost per workstream"); style_rw_cell(ws.cell(row=r, column=8))
    for col in (9, 10): style_rw_cell(ws.cell(row=r, column=col))
    row += 1

    # Layer 2 — Timeline Compression
    r = row
    ws.cell(row=r, column=1, value="Layer 2"); style_rw_cell(ws.cell(row=r, column=1), font=Font(name="Calibri", size=9, bold=True, color="C49438"))
    ws.cell(row=r, column=2, value="Timeline Compression"); style_rw_cell(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value="Team released 25% earlier"); style_rw_cell(ws.cell(row=r, column=3))
    c = ws.cell(row=r, column=4); c.value = total_rw_cost * 0.25; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    c = ws.cell(row=r, column=5); c.value = f'=D{r}*1.4'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    c = ws.cell(row=r, column=6); c.value = f'=D{r}*3'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    c = ws.cell(row=r, column=7); c.value = f'=D{r}*5'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    ws.cell(row=r, column=8, value="25% critical path compression on 18-month timeline"); style_rw_cell(ws.cell(row=r, column=8))
    for col in (9, 10): style_rw_cell(ws.cell(row=r, column=col))
    row += 1

    # Layer 3 — Team Fungibility
    r = row
    ws.cell(row=r, column=1, value="Layer 3"); style_rw_cell(ws.cell(row=r, column=1), font=Font(name="Calibri", size=9, bold=True, color="C49438"))
    ws.cell(row=r, column=2, value="Team Fungibility"); style_rw_cell(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value="A-team on every bid"); style_rw_cell(ws.cell(row=r, column=3))
    for col in range(4, 11): style_rw_cell(ws.cell(row=r, column=col))
    ws.cell(row=r, column=8, value="Same specialists cover 2-3 concurrent bids at A-team quality"); style_rw_cell(ws.cell(row=r, column=8))
    row += 1

    # Layer 4 — Decision Velocity
    r = row
    ws.cell(row=r, column=1, value="Layer 4"); style_rw_cell(ws.cell(row=r, column=1), font=Font(name="Calibri", size=9, bold=True, color="C49438"))
    ws.cell(row=r, column=2, value="Decision Velocity"); style_rw_cell(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value="Early kill avoids"); style_rw_cell(ws.cell(row=r, column=3))
    avoidable = total_rw_cost - pre_gate_cost
    c = ws.cell(row=r, column=4); c.value = avoidable; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    c = ws.cell(row=r, column=5); c.value = f'=D{r}*1.4'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    c = ws.cell(row=r, column=6); c.value = f'=D{r}*3'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    c = ws.cell(row=r, column=7); c.value = f'=D{r}*5'; style_rw_cell(c, font=RW_NUM_FONT, number_format="£#,##0")
    ws.cell(row=r, column=8, value=f"89% of bid cost — only £{pre_gate_cost:,.0f} spent before informed gate"); style_rw_cell(ws.cell(row=r, column=8))
    for col in (9, 10): style_rw_cell(ws.cell(row=r, column=col))
    row += 1

    set_col_widths(ws, [14, 22, 22, 14, 14, 14, 14, 58, 8, 8])

    return {"rd": sum(ws_days_rw.values()), "rc": total_rw_cost, "sd": 0, "sc": 0}  # Actuals via formulas


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    rows = build_all_tasks()
    print(f"Total tasks: {len(rows)}")

    # Sort by workstream order, then activity, then task ID
    ws_order_map = {ws: i for i, ws in enumerate(WS_ORDER)}
    rows.sort(key=lambda r: (
        ws_order_map.get(r["workstream"], 99),
        r["activity"],
        r["task_id"],
    ))

    wb = Workbook()
    write_all_tasks(wb, rows)
    write_summary(wb, rows)
    total_effort, total_saved, total_pct = write_effort_reduction(wb, rows)
    rw_grand = write_real_world_impact(wb, rows)
    write_top_10(wb)
    write_roadmap(wb)
    write_external_data(wb)

    wb.save(str(OUTPUT))
    print(f"Written to: {OUTPUT}")

    # Summary stats
    high = sum(1 for r in rows if r["rating"] == "High")
    med = sum(1 for r in rows if r["rating"] == "Medium")
    low = sum(1 for r in rows if r["rating"] == "Low")
    print(f"High: {high}, Medium: {med}, Low: {low}")

    # Real-world summary
    if rw_grand:
        print(f"\n--- REAL-WORLD AI IMPACT (2,050 person-days, £1.64m base) ---")
        print(f"AI days saved:          {rw_grand['sd']:.0f} of {rw_grand['rd']:.0f} ({rw_grand['sd']/rw_grand['rd']*100:.1f}%)")
        print(f"AI cost saving (base):  £{rw_grand['sc']:,.0f}")
        print(f"AI cost saving (1.4x):  £{rw_grand['sc'] * 1.4:,.0f}")

    # Effort reduction summary (methodology baseline)
    print(f"\n--- EFFORT REDUCTION SUMMARY (methodology baseline) ---")
    print(f"Total bid effort:       {total_effort} person-days")
    print(f"AI days saved:          {total_saved} person-days")
    print(f"Remaining effort:       {round(total_effort - total_saved, 1)} person-days")
    print(f"Overall reduction:      {total_pct}%")

    # Per-workstream summary
    print(f"\nPer workstream:")
    for wscode in WS_ORDER:
        ws_activities = set(r["activity"] for r in rows if r["workstream"] == wscode)
        ws_effort = sum(ACTIVITY_EFFORT.get(a, {}).get("effortDays", 0) for a in ws_activities)
        ws_saved = round(sum(r["ai_days_saved"] for r in rows if r["workstream"] == wscode), 1)
        ws_pct = round(ws_saved / max(ws_effort, 0.1) * 100, 1) if ws_effort > 0 else 0
        print(f"  {wscode:6s}  {ws_effort:5.0f} days  →  {ws_saved:5.1f} saved  ({ws_pct}%)")


if __name__ == "__main__":
    main()

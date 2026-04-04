#!/usr/bin/env python3
"""Parse methodology_mapping_review.html and generate an Excel workbook for review input."""

import re
from html.parser import HTMLParser
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Parse HTML ───────────────────────────────────────────────────────────────

def strip_tags(html_str):
    """Remove HTML tags, decode entities, collapse whitespace."""
    text = re.sub(r'<[^>]+>', ' ', html_str)
    text = text.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
    text = text.replace('&ndash;', '–').replace('&mdash;', '—').replace('&nbsp;', ' ')
    text = text.replace('&#8217;', "'").replace('&#8220;', '"').replace('&#8221;', '"')
    return re.sub(r'\s+', ' ', text).strip()


def parse_methodology(html):
    """Regex-based parser for the methodology mapping review HTML."""
    workstreams = []
    gaps = []

    # Split into workstream blocks using <!-- WS-CODE --> comments or ws-hdr divs
    ws_blocks = re.split(r'(?=<div class="ws-hdr)', html)

    for block in ws_blocks:
        # Must contain a workstream header
        ws_m = re.search(r'<div class="ws-code">(.*?)</div>', block)
        if not ws_m:
            # Check for gaps section in non-workstream blocks
            gap_matches = re.finditer(
                r'<div class="gap[^"]*gap-(missing|extra|partial)[^"]*">(.*?)</div>\s*</div>',
                block, re.DOTALL
            )
            for gm in gap_matches:
                gap_type = {'missing': 'MISSING', 'extra': 'EXTENSION', 'partial': 'PARTIAL'}[gm.group(1)]
                gap_html = gm.group(2)
                title = strip_tags(m.group(1)) if (m := re.search(r'<div class="gap-title">(.*?)</div>', gap_html)) else ''
                desc = strip_tags(m.group(1)) if (m := re.search(r'<div class="gap-desc">(.*?)</div>', gap_html)) else ''
                action = strip_tags(m.group(1)) if (m := re.search(r'<div class="gap-action">(.*?)</div>', gap_html)) else ''
                gaps.append({'type': gap_type, 'title': title, 'desc': desc, 'action': action})
            continue

        ws_code_raw = strip_tags(ws_m.group(1))
        ws_code = ws_code_raw.split('—')[0].strip() if '—' in ws_code_raw else ws_code_raw.split('–')[0].strip()
        ws_name = strip_tags(m.group(1)) if (m := re.search(r'<div class="ws-name">(.*?)</div>', block)) else ''
        ws_desc = strip_tags(m.group(1)) if (m := re.search(r'<div class="ws-desc">(.*?)</div>', block)) else ''

        ws = {'code': ws_code, 'name': ws_name, 'desc': ws_desc, 'activities': []}
        workstreams.append(ws)

        # Split into activity blocks
        act_blocks = re.split(r'(?=<div class="act">)', block)
        for ablock in act_blocks:
            act_m = re.search(r'<div class="act-code">(.*?)</div>', ablock)
            if not act_m:
                continue

            act_code = strip_tags(act_m.group(1))
            act_name = strip_tags(m.group(1)) if (m := re.search(r'<div class="act-name">(.*?)</div>', ablock)) else ''
            act_output = strip_tags(m.group(1)).replace('Output:', '').strip() if (m := re.search(r'<div class="act-output">(.*?)</div>', ablock)) else ''

            src_m = re.search(r'<div class="act-src[^"]*src-(\w+)[^"]*">(.*?)</div>', ablock)
            act_src = strip_tags(src_m.group(2)) if src_m else ''

            act = {'code': act_code, 'name': act_name, 'src': act_src, 'output': act_output, 'l2s': [], 'notes': []}
            ws['activities'].append(act)

            # Extract review notes
            for note_m in re.finditer(r'<div class="note">(.*?)</div>\s*</div>', ablock, re.DOTALL):
                note_text = strip_tags(note_m.group(1)).replace('Review Note', '').strip()
                if note_text:
                    act['notes'].append(note_text)

            # Extract L2 blocks
            l2_blocks = re.split(r'(?=<div class="l2">)', ablock)
            for l2block in l2_blocks:
                l2_m = re.search(r'<div class="l2-name">(.*?)</div>', l2block, re.DOTALL)
                if not l2_m:
                    continue

                l2_raw = strip_tags(l2_m.group(1))
                parts = l2_raw.split('←')
                l2_name = parts[0].replace('L2:', '').strip()
                l2_src = parts[1].strip() if len(parts) > 1 else ''

                l2 = {'name': l2_name, 'src': l2_src, 'tasks': []}
                act['l2s'].append(l2)

                # Extract L3 task rows from tables
                for table_m in re.finditer(r'<table class="l3-tbl">(.*?)</table>', l2block, re.DOTALL):
                    rows = re.findall(r'<tr>(.*?)</tr>', table_m.group(1), re.DOTALL)
                    for row_html in rows:
                        # Skip header rows
                        if '<th' in row_html:
                            continue
                        cells = [strip_tags(c) for c in re.findall(r'<td[^>]*>(.*?)</td>', row_html, re.DOTALL)]
                        if len(cells) >= 7:
                            task_name = cells[0]
                            if task_name.startswith('⚠') or task_name.startswith('\u26a0'):
                                act['notes'].append(task_name)
                                continue
                            task = {
                                'name': task_name,
                                'r': cells[1],
                                'a': cells[2],
                                'c': cells[3],
                                'i': cells[4],
                                'inputs': cells[5],
                                'outputs': cells[6],
                                'effort': cells[7] if len(cells) > 7 else '',
                                'type': cells[8] if len(cells) > 8 else '',
                            }
                            l2['tasks'].append(task)

    # Parse all gap cards — extract the full block between gap open and close
    gap_starts = [(m.start(), m.group(1)) for m in re.finditer(r'<div class="gap\s+gap-(missing|extra|partial)">', html)]
    for start_pos, gap_key in gap_starts:
        # Walk forward counting div depth to find matching </div>
        depth = 0
        i = start_pos
        end_pos = len(html)
        while i < len(html):
            if html[i:i+4] == '<div':
                depth += 1
                i += 4
            elif html[i:i+6] == '</div>':
                depth -= 1
                if depth == 0:
                    end_pos = i + 6
                    break
                i += 6
            else:
                i += 1
        gap_html = html[start_pos:end_pos]

        gap_type = {'missing': 'MISSING', 'extra': 'EXTENSION', 'partial': 'PARTIAL'}[gap_key]
        title = ''
        if (m := re.search(r'<div class="gap-title">(.*?)</div>', gap_html)):
            title = strip_tags(m.group(1))
        elif (m := re.search(r'<div class="gap-label">(.*?)</div>', gap_html)):
            title = strip_tags(m.group(1))
        desc = strip_tags(m.group(1)) if (m := re.search(r'<div class="gap-desc">(.*?)</div>', gap_html, re.DOTALL)) else ''
        action = strip_tags(m.group(1)) if (m := re.search(r'<div class="gap-action">(.*?)</div>', gap_html, re.DOTALL)) else ''
        gaps.append({'type': gap_type, 'title': title, 'desc': desc, 'action': action})

    # Parse gate comparison table
    gate_rows = []
    gate_table = re.search(r'<table class="gate-tbl">(.*?)</table>', html, re.DOTALL)
    if gate_table:
        rows = re.findall(r'<tr>(.*?)</tr>', gate_table.group(1), re.DOTALL)
        for row_html in rows:
            if '<th' in row_html:
                continue
            cells = [strip_tags(c) for c in re.findall(r'<td[^>]*>(.*?)</td>', row_html, re.DOTALL)]
            if len(cells) >= 5:
                gate_rows.append(cells)

    return workstreams, gaps, gate_rows


# ── Generate Excel ───────────────────────────────────────────────────────────

class ParseResult:
    """Simple container to hold parsed data."""
    def __init__(self, workstreams, gaps, gate_rows=None):
        self.workstreams = workstreams
        self.gaps = gaps
        self.gate_rows = gate_rows or []


def build_workbook(parser):
    wb = Workbook()

    # Colours
    navy = '0F1724'
    dark_bg = '162032'
    gold = 'C49438'
    white = 'E8ECF1'
    grey = '94A3B8'
    green = '34D399'
    amber = 'FBBF24'
    blue = '60A5FA'
    purple = 'A78BFA'

    # Fonts
    hdr_font = Font(name='Calibri', bold=True, size=11, color=white)
    ws_font = Font(name='Calibri', bold=True, size=12, color=gold)
    act_font = Font(name='Calibri', bold=True, size=11, color=white)
    data_font = Font(name='Calibri', size=10, color='333333')
    note_font = Font(name='Calibri', size=10, color=gold, italic=True)
    review_font = Font(name='Calibri', size=10, color='CC0000', bold=True)

    # Fills
    hdr_fill = PatternFill('solid', fgColor=navy)
    ws_fill = PatternFill('solid', fgColor='1D2B40')
    act_fill = PatternFill('solid', fgColor='25354E')
    light_fill = PatternFill('solid', fgColor='F7F8FA')
    review_fill = PatternFill('solid', fgColor='FFF3CD')

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    wrap = Alignment(wrap_text=True, vertical='top')

    # ── Sheet 1: L3 Task Review ──────────────────────────────────────────

    ws = wb.active
    ws.title = 'L3 Task Review'

    # Column definitions
    cols = [
        ('Workstream', 12),
        ('Activity Code', 13),
        ('Activity Name', 30),
        ('Mapping Status', 14),
        ('Activity Output', 30),
        ('L2 Sub-Process', 30),
        ('Playbook Ref', 15),
        ('L3 Task', 45),
        ('R (Responsible)', 18),
        ('A (Accountable)', 18),
        ('C (Consulted)', 20),
        ('I (Informed)', 18),
        ('Inputs (pre-requisites)', 35),
        ('Outputs (deliverables)', 35),
        ('Effort', 10),
        ('Type', 12),
        ('YOUR REVIEW: Accept? (Y/N/Modify)', 20),
        ('YOUR REVIEW: Comments / Changes', 40),
    ]

    # Header row
    for col_idx, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        if col_idx >= 17:
            cell.font = Font(name='Calibri', bold=True, size=11, color='CC0000')
            cell.fill = PatternFill('solid', fgColor='FFF3CD')
        else:
            cell.font = hdr_font
            cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'
    ws.freeze_panes = 'A2'

    row_num = 2
    for workstream in parser.workstreams:
        for act in workstream['activities']:
            if not act['l2s']:
                # Activity with no L3 tasks yet - still needs a row
                cell_data = [
                    workstream['code'],
                    act['code'],
                    act['name'],
                    act['src'],
                    act['output'],
                    '',  # no L2
                    '',  # no ref
                    '(No L3 tasks defined - needs authoring)',
                    '', '', '', '', '', '', '', '',
                    '',  # review accept
                    '',  # review comment
                ]
                for col_idx, val in enumerate(cell_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.font = data_font
                    cell.alignment = wrap
                    cell.border = thin_border
                    if col_idx >= 17:
                        cell.fill = review_fill
                row_num += 1
            else:
                for l2 in act['l2s']:
                    if not l2['tasks']:
                        cell_data = [
                            workstream['code'],
                            act['code'],
                            act['name'],
                            act['src'],
                            act['output'],
                            l2['name'],
                            l2['src'],
                            '(No L3 tasks defined - needs authoring)',
                            '', '', '', '', '', '', '', '',
                            '', '',
                        ]
                        for col_idx, val in enumerate(cell_data, 1):
                            cell = ws.cell(row=row_num, column=col_idx, value=val)
                            cell.font = data_font
                            cell.alignment = wrap
                            cell.border = thin_border
                            if col_idx >= 17:
                                cell.fill = review_fill
                        row_num += 1
                    else:
                        for task in l2['tasks']:
                            cell_data = [
                                workstream['code'],
                                act['code'],
                                act['name'],
                                act['src'],
                                act['output'],
                                l2['name'],
                                l2['src'],
                                task['name'],
                                task['r'],
                                task['a'],
                                task['c'],
                                task['i'],
                                task['inputs'],
                                task['outputs'],
                                task['effort'],
                                task['type'],
                                '',  # review accept
                                '',  # review comment
                            ]
                            for col_idx, val in enumerate(cell_data, 1):
                                cell = ws.cell(row=row_num, column=col_idx, value=val)
                                cell.font = data_font
                                cell.alignment = wrap
                                cell.border = thin_border
                                if col_idx >= 17:
                                    cell.fill = review_fill
                            row_num += 1

            # Add review notes as separate rows
            for note in act.get('notes', []):
                ws.cell(row=row_num, column=1, value=workstream['code']).font = data_font
                ws.cell(row=row_num, column=2, value=act['code']).font = data_font
                note_cell = ws.cell(row=row_num, column=8, value=f'REVIEW NOTE: {note}')
                note_cell.font = note_font
                note_cell.alignment = wrap
                for col_idx in range(1, len(cols) + 1):
                    ws.cell(row=row_num, column=col_idx).border = thin_border
                    ws.cell(row=row_num, column=col_idx).fill = PatternFill('solid', fgColor='FFF8E7')
                    if col_idx >= 17:
                        ws.cell(row=row_num, column=col_idx).fill = review_fill
                row_num += 1

    # ── Sheet 2: Gaps & Extensions ───────────────────────────────────────

    ws2 = wb.create_sheet('Gaps & Extensions')
    gap_cols = [
        ('Type', 14),
        ('Title', 40),
        ('Description', 60),
        ('Suggested Action', 50),
        ('YOUR REVIEW: Accept? (Y/N)', 18),
        ('YOUR REVIEW: Comments', 40),
    ]

    for col_idx, (name, width) in enumerate(gap_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=name)
        if col_idx >= 5:
            cell.font = Font(name='Calibri', bold=True, size=11, color='CC0000')
            cell.fill = PatternFill('solid', fgColor='FFF3CD')
        else:
            cell.font = hdr_font
            cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.auto_filter.ref = f'A1:{get_column_letter(len(gap_cols))}1'
    ws2.freeze_panes = 'A2'

    for r_idx, gap in enumerate(parser.gaps, 2):
        vals = [gap['type'], gap['title'], gap['desc'], gap['action'], '', '']
        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=r_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = wrap
            cell.border = thin_border
            if col_idx >= 5:
                cell.fill = review_fill

    # ── Sheet 3: Gate Comparison ───────────────────────────────────────

    if parser.gate_rows:
        ws3g = wb.create_sheet('Gate Comparison')
        gate_cols = [
            ('Playbook Gate', 30),
            ('Phase', 14),
            ('Product Gate', 25),
            ('Match', 14),
            ('Notes', 55),
            ('YOUR REVIEW: Accept? (Y/N)', 18),
            ('YOUR REVIEW: Comments', 40),
        ]
        for col_idx, (name, width) in enumerate(gate_cols, 1):
            cell = ws3g.cell(row=1, column=col_idx, value=name)
            if col_idx >= 6:
                cell.font = Font(name='Calibri', bold=True, size=11, color='CC0000')
                cell.fill = PatternFill('solid', fgColor='FFF3CD')
            else:
                cell.font = hdr_font
                cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=True, vertical='bottom')
            ws3g.column_dimensions[get_column_letter(col_idx)].width = width

        ws3g.auto_filter.ref = f'A1:{get_column_letter(len(gate_cols))}1'
        ws3g.freeze_panes = 'A2'

        for r_idx, row in enumerate(parser.gate_rows, 2):
            vals = row[:5] + ['', '']
            for col_idx, val in enumerate(vals, 1):
                cell = ws3g.cell(row=r_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = wrap
                cell.border = thin_border
                if col_idx >= 6:
                    cell.fill = review_fill

    # ── Sheet 4: Instructions ────────────────────────────────────────────

    ws3 = wb.create_sheet('Instructions')
    ws3.column_dimensions['A'].width = 80

    instructions = [
        'PWIN Methodology Mapping Review — How To Use This Workbook',
        '',
        'This workbook is extracted from methodology_mapping_review.html for your line-by-line review.',
        '',
        'SHEET 1: L3 Task Review',
        '  - Each row is one L3 task (the actual activity someone performs)',
        '  - Columns A-P are the current proposed content — read but do not edit these',
        '  - Column Q (Accept?): Enter Y to accept, N to reject, or M to modify',
        '  - Column R (Comments/Changes): Write your changes, corrections, or new task descriptions here',
        '  - Yellow rows with "REVIEW NOTE:" are questions or flags for your attention — respond in Column R',
        '  - Rows saying "(No L3 tasks defined)" need you to write the tasks in Column R',
        '',
        'SHEET 2: Gaps & Extensions',
        '  - Playbook content missing from the product, or product content not in the playbook',
        '  - Column E (Accept?): Y to include, N to exclude',
        '  - Column F (Comments): Any changes or reasoning',
        '',
        'COLUMN DEFINITIONS:',
        '  Workstream — The workstream code (SAL, SOL, COM, etc.)',
        '  Activity Code — Unique ID for the L1 activity (e.g. SAL-01)',
        '  Activity Name — What the activity is called',
        '  Mapping Status — MAPPED / PARTIAL / PRODUCT ONLY / NEW',
        '  Activity Output — The primary deliverable of the whole activity',
        '  L2 Sub-Process — The grouping within the activity',
        '  Playbook Ref — Which playbook section this maps to',
        '  L3 Task — The specific task an individual performs (this is what you are reviewing)',
        '  R (Responsible) — Who does the work',
        '  A (Accountable) — Who approves / is ultimately answerable',
        '  C (Consulted) — Who provides input before the task',
        '  I (Informed) — Who is told after the task',
        '  Inputs — What you need before you can start this task (pre-requisites, source documents)',
        '  Outputs — What this task produces (named deliverable / artifact)',
        '  Effort — Relative effort: Low / Medium / High',
        '  Type — Execution pattern: Sequential / Parallel / Iterative',
        '',
        'TIPS:',
        '  - Use Excel filters on Column A (Workstream) to review one workstream at a time',
        '  - Use Column D (Mapping Status) filter to focus on PRODUCT ONLY or NEW items first',
        '  - If you want to add entirely new L3 tasks, add new rows below the relevant activity',
        '  - If you want to reorder tasks, note the desired order in the Comments column',
        '  - Save as .xlsx and upload back — Claude will parse your review columns automatically',
    ]

    for r_idx, line in enumerate(instructions, 1):
        cell = ws3.cell(row=r_idx, column=1, value=line)
        if r_idx == 1:
            cell.font = Font(name='Calibri', bold=True, size=14)
        elif line.startswith('SHEET') or line.startswith('COLUMN') or line.startswith('TIPS'):
            cell.font = Font(name='Calibri', bold=True, size=11)
        else:
            cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(wrap_text=True)

    return wb


# ── Main ─────────────────────────────────────────────────────────────────────

html_path = Path(__file__).parent / 'methodology_mapping_review.html'
html = html_path.read_text(encoding='utf-8')

workstreams, gaps, gate_rows = parse_methodology(html)
parser = ParseResult(workstreams, gaps, gate_rows)

wb = build_workbook(parser)
out_path = Path(__file__).parent / 'methodology_mapping_review.xlsx'
wb.save(out_path)

# Summary
total_tasks = sum(
    len(t['tasks'])
    for ws in parser.workstreams
    for act in ws['activities']
    for t in act['l2s']
)
total_notes = sum(len(act['notes']) for ws in parser.workstreams for act in ws['activities'])
print(f'Workstreams: {len(parser.workstreams)}')
print(f'Activities:  {sum(len(ws["activities"]) for ws in parser.workstreams)}')
print(f'L3 Tasks:    {total_tasks}')
print(f'Review Notes: {total_notes}')
print(f'Gaps:        {len(parser.gaps)}')
print(f'Saved to:    {out_path}')
